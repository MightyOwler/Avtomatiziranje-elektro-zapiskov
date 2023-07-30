import re
from datetime import datetime
import csv
from Meritev import pretvori_v_osnovne_enote, SEZNAM_VRST_MERITEV
import os
from openpyxl import load_workbook


PRAZNO = " "

st_vnesenih_meritev = 0
st_vnesenih_meritev_RCD = 0


def pretvori_string_milisekund_v_ustrezen_format(string):
    return string.replace(",", ".").replace(" ms", "").replace(">", "")


def vrednoti_string(s):
    """
    Funkcija, ki nam omogoča da razvrstimo meritve po velikosti kljub znaku >
    (to je konkretno pomembno pri RISO in RLOW4)
    """
    if ">" in s:
        return 1000000000000
    if "X" in s:
        return 0
    else:
        return pretvori_v_osnovne_enote(s.replace(",", "."))


def zapisi_kocko_meritev_v_excel_instalacije(
    kocka, loceno_besedilo, slovar_kock_in_ustreznih_poti
):
    """
    Zapiše meritev v csv datoteko, ki je primerna za obdelavo v csv-ju

    Args: kocka, loceno_besedilo, slovar_kock_in_ustreznih_poti
    """

    CSVFILE_RLOW4 = os.path.join(
        "Csvji", "Instalacije", "csv_za_excel_datoteko_RLOW4.csv"
    )
    CSVFILE_VARISTOR = os.path.join(
        "Csvji", "Instalacije", "csv_za_excel_datoteko_VARISTOR.csv"
    )
    CSVFILE_OSNOVNE = os.path.join(
        "Csvji", "Instalacije", "csv_za_excel_datoteko_osnovne.csv"
    )
    CSVFILE_RCD = os.path.join("Csvji", "Instalacije", "csv_za_excel_datoteko_RCD.csv")

    global st_vnesenih_meritev
    global st_vnesenih_meritev_RCD

    ipsc_vrednosti_zloop4w = 10**18
    ipsc_vrednosti_zline4w = 10**18
    ustrezna_meritev_zloop4w = None
    ustrezna_meritev_zline4w = None

    komentar = ""
    (
        uln,
        zln,
        ipsc_ln,
        ipsc_lpe,
        rlpe,
        dU,
        zlpe,
        glavna_izenac_povezava,
        I_dN,
        t1x,
        t5x,
        Uc,
        ia_psc_navidezni_stolpec,
        maxRplusRminus,
        tip_varovalke,
        I_varovalke,
        t_varovalke,
        isc_faktor,
        ime,
    ) = ("X" for _ in range(19))

    pot = (
        slovar_kock_in_ustreznih_poti[loceno_besedilo.index(kocka)]
        .replace("\n", " ")
        .strip()
    )
    pot_locena_na_elemente = pot.replace("\n", " ").strip().split("//")
    for element in pot_locena_na_elemente:
        if "Imenovanje: " in element.strip():
            ime = element.replace("Imenovanje: ", "")
            if "Circuit F" in ime:
                ime = ime.replace("Circuit ", "")
            elif re.search(r"Circuit\d", ime):
                ime = ime.replace("Circuit", "F")
    if not ime:
        ime = "X"

    # Za vsak slučaj preverimo, ali pot ne obstaja
    if pot is None:
        pot = "X"
        print("Napaka: Ni poti v slovarju poti", slovar_kock_in_ustreznih_poti)

    vrste_meritev_v_kocki = [meritev.doloci_vrsto_meritve() for meritev in kocka]
    slovar_vrst_meritev = {
        i: vrste_meritev_v_kocki.count(i) for i in SEZNAM_VRST_MERITEV
    }

    # Najprej pogleda za R iso

    if slovar_vrst_meritev["R iso"] + slovar_vrst_meritev["R IZO"] > 1:
        nova_kocka = []
        seznam_riso_meritev = []
        for meritev in kocka:
            if meritev.doloci_vrsto_meritve() in ["R iso", "R IZO"]:
                # inšturment vedno izpiše v MΩ
                seznam_riso_meritev.append(
                    meritev.najdi_Rlpe().replace(" MΩ", "").replace(",", ".")
                )
        seznam_riso_meritev.sort(key=lambda x: vrednoti_string(x))
        riso_meritev_z_minimalno = seznam_riso_meritev[0]
        for idx, meritev in enumerate(kocka):
            if meritev.doloci_vrsto_meritve() in ["R iso", "R IZO"]:
                if idx == seznam_riso_meritev.index(riso_meritev_z_minimalno):
                    nova_kocka.append(meritev)
            else:
                nova_kocka.append(meritev)

        kocka = nova_kocka

    if slovar_vrst_meritev["R low 4"] > 0:
        with open(CSVFILE_RLOW4, "a", encoding="utf-8", newline="") as csvfile:
            writer = csv.writer(
                csvfile, delimiter=";", quotechar='"', quoting=csv.QUOTE_MINIMAL
            )
            seznam_Rjev_za_rlow4_meritve = []

            # Spodnji loop se da močno izboljšati, veliko stvari je nekoliko nenavadnih

            for meritev in kocka:
                if meritev.doloci_vrsto_meritve() == "R low 4":
                    komentar = meritev.najdi_komentar()
                    seznam_Rjev_za_rlow4_meritve.append(meritev.najdi_R())
                    if (
                        # >1999 je zgornja meja, nastavljena na napravi
                        ">1999" in meritev.najdi_R_pozitivno()
                        or ">1999" in meritev.najdi_R_negativno()
                    ):
                        maxRplusRminus = ">1999"
                    else:
                        R_pozitivno_int = (
                            # TODO ali se res ne da lepše manipulirati s temi številkami?
                            meritev.najdi_R_pozitivno()
                            .replace(",", ".")
                            .replace(" Ω", "")
                            .replace(">", "")
                        )
                        if R_pozitivno_int != "X":
                            R_pozitivno_int = float(R_pozitivno_int)

                        R_negativno_int = (
                            meritev.najdi_R_negativno()
                            .replace(",", ".")
                            .replace(" Ω", "")
                            .replace(">", "")
                        )
                        if R_negativno_int != "X":
                            R_negativno_int = float(R_negativno_int)

                        # Pri kateri velikosti spremenimo v int?
                        if R_negativno_int == "X" or R_pozitivno_int == "X":
                            maxRplusRminus = "X"
                        else:
                            if max(R_negativno_int, R_pozitivno_int) >= 100:
                                maxRplusRminus = (
                                    f"{int(max(R_negativno_int, R_pozitivno_int))}"
                                )
                            else:
                                maxRplusRminus = (
                                    f"{max(R_negativno_int, R_pozitivno_int)}"
                                )
                    R = meritev.najdi_R()
                    vrsta_meritve = meritev.doloci_vrsto_meritve()
                    array_ki_ga_zapisemo_v_csv = [
                        ime,
                        R,
                        maxRplusRminus,
                        komentar,
                        vrsta_meritve,
                        pot,
                    ]
                    writer.writerow(array_ki_ga_zapisemo_v_csv)
            csvfile.close()
            seznam_Rjev_za_rlow4_meritve.sort(key=lambda x: float(vrednoti_string(x)))
            rlow4_meritev_z_minimalno = seznam_Rjev_za_rlow4_meritve[0]

        if slovar_vrst_meritev["RCD Auto"] > 1:
            print("Napaka: Imamo 2 ali več RCD Auto meritvi v eni kocki!")

        for idx, vrednost_meritve in enumerate(seznam_Rjev_za_rlow4_meritve):
            if idx == seznam_Rjev_za_rlow4_meritve.index(rlow4_meritev_z_minimalno):
                # Glavno izenačitveno povezavo potrebujemo samo pri eni meritvi
                glavna_izenac_povezava = vrednost_meritve
                break

    for meritev in kocka:
        komentar = meritev.najdi_komentar()
        vrsta_meritve = meritev.doloci_vrsto_meritve()
        if vrsta_meritve in ["R iso", "R IZO"]:
            if vrsta_meritve == "R iso":
                rlpe = meritev.najdi_Rlpe().replace(" MΩ", "")
            else:
                rlpe = meritev.najdi_Riso().replace(" MΩ", "")

        if vrsta_meritve == "Padec napetosti":
            dU = meritev.najdi_dU()

        if vrsta_meritve == "RCD Auto":
            Uc = meritev.najdi_Uc()
            I_dN = meritev.najdi_I_dN()
            t1x_plus = pretvori_string_milisekund_v_ustrezen_format(
                meritev.najdi_t_IΔN_x1_plus()
            )
            t1x_neg = pretvori_string_milisekund_v_ustrezen_format(
                meritev.najdi_t_IΔN_x1_minus()
            )
            t5x_plus = pretvori_string_milisekund_v_ustrezen_format(
                meritev.najdi_t_IΔN_x5_plus()
            )
            t5x_neg = pretvori_string_milisekund_v_ustrezen_format(
                meritev.najdi_t_IΔN_x5_minus()
            )

            # V primeru, da vse količine obstajajo
            if "X" not in [t1x_plus, t1x_neg, t5x_plus, t5x_neg]:
                t1x_plus = float(t1x_plus)
                t1x_neg = float(t1x_neg)
                t5x_plus = float(t5x_plus)
                t5x_neg = float(t5x_neg)
                max_t1 = max(t1x_plus, t1x_neg)
                max_t5 = max(t5x_plus, t5x_neg)

                # Pri vrednosti večji ali enaki 100 rezultat zaokrožimo
                if max_t1 >= 100:
                    t1x = f"{int(max_t1)}".replace(".", ",")
                else:
                    t1x = f"{max_t1}".replace(".", ",")

                if max_t5 >= 100:
                    t5x = f"{int(max_t5)}".replace(".", ",")
                else:
                    t5x = f"{max_t5}".replace(".", ",")
            else:
                t1x = "X"
                t5x = "X"

        if vrsta_meritve == "Varistor":
            with open(
                CSVFILE_VARISTOR,
                "a",
                encoding="utf-8",
                newline="",
            ) as csvfile:
                writer = csv.writer(
                    csvfile, delimiter=";", quotechar='"', quoting=csv.QUOTE_MINIMAL
                )

                uac = meritev.najdi_Uac()
                udc = meritev.najdi_Udc()

                array_ki_ga_zapisemo_v_csv = [
                    ime,
                    uac,
                    udc,
                    komentar,
                    vrsta_meritve,
                    pot,
                ]
                writer.writerow(array_ki_ga_zapisemo_v_csv)
                csvfile.close()

    # Stvar je treba narediti v večih korakih, saj lahko podatki v posamezni meritvi vplivajo na celotno kocko ali kasnejše.
    # najprej odpravimo AUTO TN

    for meritev in kocka:
        vrsta_meritve = meritev.doloci_vrsto_meritve()

        if vrsta_meritve == "AUTO TN":
            with open(
                CSVFILE_OSNOVNE,
                "a",
                encoding="utf-8",
                newline="",
            ) as csvfile:
                writer = csv.writer(
                    csvfile, delimiter=";", quotechar='"', quoting=csv.QUOTE_MINIMAL
                )

                # Tole je treba izboljšati
                glavna_izenac_povezava = meritev.najdi_R()
                uln = meritev.najdi_Uln()
                zln = meritev.najdi_Z_LN()
                ipsc_ln = meritev.najdi_Ipsc_LN()
                ipsc_lpe = meritev.najdi_Ipsc_LPE()
                dU = meritev.najdi_dU()
                zlpe = meritev.najdi_Z_LPE()
                ia_psc_navidezni_stolpec = meritev.najdi_Ia_Ipsc()
                tip_varovalke = meritev.najdi_tip_varovalke()
                I_varovalke = meritev.najdi_I_varovalke()
                t_varovalke = meritev.najdi_t_varovalke()
                isc_faktor = meritev.najdi_Isc_faktor()
                komentar = meritev.najdi_komentar()

                st_vnesenih_meritev += 1
                array_ki_ga_zapisemo_v_csv = [
                    st_vnesenih_meritev,
                    ime,
                    PRAZNO,
                    PRAZNO,
                    PRAZNO,
                    glavna_izenac_povezava,
                    PRAZNO,
                    rlpe,
                    tip_varovalke,
                    I_varovalke,
                    t_varovalke,
                    f"{zlpe}/{ipsc_ln}",
                    f"{zln}/{ipsc_lpe}/{dU}",
                    PRAZNO,
                    I_dN,
                    PRAZNO,
                    t1x,
                    t5x,
                    Uc,
                    PRAZNO,
                    komentar,
                    vrsta_meritve,
                    uln,
                    maxRplusRminus,
                    isc_faktor,
                    ia_psc_navidezni_stolpec,
                    pot,
                ]
                writer.writerow(array_ki_ga_zapisemo_v_csv)
                csvfile.close()

        if vrsta_meritve in ["ZLOOP 4W", "Zline 4W"]:
            if vrsta_meritve == "ZLOOP 4W":
                # print(meritev.najdi_Ipsc())
                if meritev.najdi_Ipsc() == "X":
                    ustrezna_meritev_zloop4w = meritev
                    ipsc_vrednosti_zloop4w = "X"
                elif float(meritev.najdi_Ipsc()) < ipsc_vrednosti_zloop4w:
                    ustrezna_meritev_zloop4w = meritev
                    ipsc_vrednosti_zloop4w = float(meritev.najdi_Ipsc())
            else:
                if float(meritev.najdi_Ipsc()) < ipsc_vrednosti_zline4w:
                    ustrezna_meritev_zline4w = meritev
                    ipsc_vrednosti_zline4w = float(meritev.najdi_Ipsc())

    if "ZLOOP 4W" in vrste_meritev_v_kocki or "ZLINE 4W" in vrste_meritev_v_kocki:
        with open(
            CSVFILE_OSNOVNE,
            "a",
            encoding="utf-8",
            newline="",
        ) as csvfile:
            writer = csv.writer(
                csvfile, delimiter=";", quotechar='"', quoting=csv.QUOTE_MINIMAL
            )

            if (
                not ustrezna_meritev_zline4w
                or ustrezna_meritev_zline4w.besedilo.count("p//") > 0
            ):
                ipsc_zline, z_zline = "X", "X"
            else:
                ipsc_zline = ustrezna_meritev_zline4w.najdi_Ipsc()
                z_zline = ustrezna_meritev_zline4w.najdi_Z()

            if (
                not ustrezna_meritev_zloop4w
                or ustrezna_meritev_zloop4w.besedilo.count("p//") > 0
            ):
                (
                    ipsc_zloop,
                    z_zloop,
                    uln,
                    ipsc_lpe,
                    zlpe,
                    ia_psc_navidezni_stolpec,
                    tip_varovalke,
                    I_varovalke,
                    t_varovalke,
                    isc_faktor,
                    komentar,
                ) = ("X" for _ in range(11))

            else:
                ipsc_zloop = ustrezna_meritev_zloop4w.najdi_Ipsc()
                z_zloop = ustrezna_meritev_zloop4w.najdi_Z()
                uln = ustrezna_meritev_zloop4w.najdi_Uln()
                ipsc_lpe = ustrezna_meritev_zloop4w.najdi_Ipsc_LPE()
                zlpe = ustrezna_meritev_zloop4w.najdi_Z_LPE()
                ia_psc_navidezni_stolpec = ustrezna_meritev_zloop4w.najdi_Ia_Ipsc()
                tip_varovalke = ustrezna_meritev_zloop4w.najdi_tip_varovalke()
                I_varovalke = ustrezna_meritev_zloop4w.najdi_I_varovalke()
                t_varovalke = ustrezna_meritev_zloop4w.najdi_t_varovalke()
                isc_faktor = ustrezna_meritev_zloop4w.najdi_Isc_faktor()
                vrsta_meritve = "ZLOOP 4W / ZLINE 4W"
                komentar = ustrezna_meritev_zloop4w.najdi_komentar()

                if not dU:
                    dU = "X"

                st_vnesenih_meritev += 1
                array_ki_ga_zapisemo_v_csv = [
                    st_vnesenih_meritev,
                    ime,
                    PRAZNO,
                    PRAZNO,
                    PRAZNO,
                    glavna_izenac_povezava,
                    PRAZNO,
                    rlpe,
                    tip_varovalke,
                    I_varovalke,
                    t_varovalke,
                    f"{z_zloop}/{ipsc_zloop}",
                    f"{z_zline}/{ipsc_zline}/{dU}",
                    PRAZNO,
                    I_dN,
                    PRAZNO,
                    t1x,
                    t5x,
                    Uc,
                    PRAZNO,
                    komentar,
                    vrsta_meritve,
                    uln,
                    maxRplusRminus,
                    isc_faktor,
                    ia_psc_navidezni_stolpec,
                    pot,
                ]
                writer.writerow(array_ki_ga_zapisemo_v_csv)
                csvfile.close()

    # nato odpravimo Zloop / Zine

    if "Zloop" in vrste_meritev_v_kocki or "Z LINE" in vrste_meritev_v_kocki:
        ustrezni_zline_3 = []
        ustrezni_zloop_3 = []
        for meritev in kocka:
            vrsta_meritve = meritev.doloci_vrsto_meritve()
            if vrsta_meritve == "Zloop":
                ustrezni_zloop_3.append(meritev)
            if vrsta_meritve == "Z LINE":
                if "400" == meritev.najdi_Un():
                    ustrezni_zline_3.append(meritev)

        # Safety check v primeru, da ni treh zloop/zline elementov
        if (
            len(ustrezni_zline_3) != len(ustrezni_zloop_3)
            or len(ustrezni_zline_3) != 3
            or len(ustrezni_zloop_3) != 3
        ):
            print("\nNapaka: Nekaj ni v redu s številom zloop/zlinov")
            print("Pot problematične meritve:", pot.strip())
            print("Dolžina zloop", len(ustrezni_zloop_3))
            print("Dolžina zline", len(ustrezni_zline_3))

        else:
            for i in range(3):
                with open(
                    CSVFILE_OSNOVNE,
                    "a",
                    encoding="utf-8",
                    newline="",
                ) as csvfile:
                    writer = csv.writer(
                        csvfile, delimiter=";", quotechar='"', quoting=csv.QUOTE_MINIMAL
                    )

                    if ustrezni_zline_3[i].besedilo.count("p//") > 0:
                        ipsc_zline, z_zline = "X", "X"
                    else:
                        ipsc_zline = ustrezni_zline_3[i].najdi_Ipsc()
                        z_zline = ustrezni_zline_3[i].najdi_Z()

                    if ustrezni_zloop_3[i].besedilo.count("p//") > 0:
                        ipsc_zloop, z_zloop = "X", "X"
                    else:
                        ipsc_zloop = ustrezni_zloop_3[i].najdi_Ipsc()
                        z_zloop = ustrezni_zloop_3[i].najdi_Z()

                    if ustrezni_zloop_3[i].besedilo.count("p//") > 0:
                        (
                            uln,
                            ipsc_lpe,
                            zlpe,
                            ia_psc_navidezni_stolpec,
                            tip_varovalke,
                            I_varovalke,
                            t_varovalke,
                            isc_faktor,
                            komentar,
                        ) = ("X" for _ in range(9))
                    else:
                        uln = ustrezni_zloop_3[i].najdi_Uln()
                        ipsc_lpe = ustrezni_zloop_3[i].najdi_Ipsc_LPE()
                        zlpe = ustrezni_zloop_3[i].najdi_Z_LPE()
                        ia_psc_navidezni_stolpec = ustrezni_zloop_3[i].najdi_Ia_Ipsc()
                        tip_varovalke = ustrezni_zloop_3[i].najdi_tip_varovalke()
                        I_varovalke = ustrezni_zloop_3[i].najdi_I_varovalke()
                        t_varovalke = ustrezni_zloop_3[i].najdi_t_varovalke()
                        isc_faktor = ustrezni_zloop_3[i].najdi_Isc_faktor()
                        vrsta_meritve = "ZLOOP / ZLINE"
                        komentar = ustrezni_zloop_3[i].najdi_komentar()

                        if not dU:
                            dU = "X"

                        st_vnesenih_meritev += 1
                        array_ki_ga_zapisemo_v_csv = [
                            st_vnesenih_meritev,
                            ime,
                            PRAZNO,
                            PRAZNO,
                            PRAZNO,
                            glavna_izenac_povezava,
                            PRAZNO,
                            rlpe,
                            tip_varovalke,
                            I_varovalke,
                            t_varovalke,
                            f"{z_zloop}/{ipsc_zloop}",
                            f"{z_zline}/{ipsc_zline}/{dU}",
                            PRAZNO,
                            I_dN,
                            PRAZNO,
                            t1x,
                            t5x,
                            Uc,
                            PRAZNO,
                            komentar,
                            vrsta_meritve,
                            uln,
                            maxRplusRminus,
                            isc_faktor,
                            ia_psc_navidezni_stolpec,
                            pot,
                        ]
                        writer.writerow(array_ki_ga_zapisemo_v_csv)
                        csvfile.close()

    # nato odpravimo RCD AUTO

    for meritev in kocka:
        vrsta_meritve = meritev.doloci_vrsto_meritve()
        if vrsta_meritve == "RCD Auto":
            with open(
                CSVFILE_RCD,
                "a",
                encoding="utf-8",
                newline="",
            ) as csvfile_RCD:
                writer = csv.writer(
                    csvfile_RCD, delimiter=";", quotechar='"', quoting=csv.QUOTE_MINIMAL
                )
                I_dN = meritev.najdi_I_dN()
                tip_rcd = meritev.najdi_Tip()
                Uc = meritev.najdi_Uc()
                ozemljitveni_sistem = meritev.najdi_Ozemljitveni_sistem()
                komentar = meritev.najdi_komentar()

                t1x_plus = pretvori_string_milisekund_v_ustrezen_format(
                    meritev.najdi_t_IΔN_x1_plus()
                )
                t1x_neg = pretvori_string_milisekund_v_ustrezen_format(
                    meritev.najdi_t_IΔN_x1_minus()
                )
                t5x_plus = pretvori_string_milisekund_v_ustrezen_format(
                    meritev.najdi_t_IΔN_x5_plus()
                )
                t5x_neg = pretvori_string_milisekund_v_ustrezen_format(
                    meritev.najdi_t_IΔN_x5_minus()
                )

                # V primeru, da vse količine obstajajo
                if "X" not in [t1x_plus, t1x_neg, t5x_plus, t5x_neg]:
                    t1x_plus = float(t1x_plus)
                    t1x_neg = float(t1x_neg)
                    t5x_plus = float(t5x_plus)
                    t5x_neg = float(t5x_neg)

                    # Pri kateri velikosti spremenimo v int?
                    if max(t1x_plus, t1x_neg) >= 100:
                        t1x = f"{int(max(t1x_plus, t1x_neg))}".replace(".", ",")
                    else:
                        t1x = f"{max(t1x_plus, t1x_neg)}".replace(".", ",")

                    if max(t5x_plus, t5x_neg) >= 100:
                        t5x = f"{int(max(t5x_plus, t5x_neg))}".replace(".", ",")
                    else:
                        t5x = f"{max(t5x_plus, t5x_neg)}".replace(".", ",")
                else:
                    print("\nNapaka: manjkajoči podatki t1x ali t5x pri", pot.strip())
                    t1x = "X"
                    t5x = "X"

                st_vnesenih_meritev_RCD += 1
                array_ki_ga_zapisemo_v_csv = [
                    st_vnesenih_meritev_RCD,
                    ime,
                    PRAZNO,
                    PRAZNO,
                    I_dN,
                    tip_rcd,
                    PRAZNO,
                    PRAZNO,
                    PRAZNO,
                    PRAZNO,
                    Uc,
                    PRAZNO,
                    t1x,
                    t5x,
                    ozemljitveni_sistem,
                    komentar,
                    vrsta_meritve,
                    pot,
                ]
                writer.writerow(array_ki_ga_zapisemo_v_csv)
                csvfile_RCD.close()


def zapisi_kocko_meritev_v_excel_elektricne_omare(
    kocka, loceno_besedilo, slovar_kock_in_ustreznih_poti
):
    pass


def zapisi_kocko_meritev_v_excel_stroji(
    kocka, loceno_besedilo, slovar_kock_in_ustreznih_poti
):
    CSVFILE_ZLOOP = os.path.join("Csvji", "Stroji", "csv_za_excel_datoteko_ZLOOP.csv")
    CSVFILE_RISO = os.path.join("Csvji", "Stroji", "csv_za_excel_datoteko_R ISO.csv")
    CSVFILE_DISCHARGE_TIME = os.path.join(
        "Csvji", "Stroji", "csv_za_excel_datoteko_DISCHARGE TIME.csv"
    )

    global st_vnesenih_meritev
    global st_vnesenih_meritev_RCD

    komentar = ""
    ime = "X"

    pot = (
        slovar_kock_in_ustreznih_poti[loceno_besedilo.index(kocka)]
        .replace("\n", " ")
        .strip()
    )
    pot_locena_na_elemente = pot.replace("\n", " ").strip().split("//")
    for element in pot_locena_na_elemente:
        if "Imenovanje: " in element.strip():
            ime = element.replace("Imenovanje: ", "")
            if "Circuit F" in ime:
                ime = ime.replace("Circuit ", "")
            elif re.search(r"Circuit\d", ime):
                ime = ime.replace("Circuit", "F")
    if not ime:
        ime = "X"

    if pot is None:
        pot = "X"
        print("Napaka: Ni poti v slovarju poti", slovar_kock_in_ustreznih_poti)

    vrste_meritev_v_kocki = [meritev.doloci_vrsto_meritve() for meritev in kocka]
    slovar_vrst_meritev = {
        i: vrste_meritev_v_kocki.count(i) for i in SEZNAM_VRST_MERITEV
    }

    # Kako razvejiva tukaj meritve?

    if slovar_vrst_meritev["Zloop"] > 0:
        with open(CSVFILE_ZLOOP, "a", encoding="utf-8", newline="") as csvfile:
            writer = csv.writer(
                csvfile, delimiter=";", quotechar='"', quoting=csv.QUOTE_MINIMAL
            )

            for meritev in kocka:
                if meritev.doloci_vrsto_meritve() == "Zloop":
                    komentar = meritev.najdi_komentar()

                    # dodaj specifične atribute

                    tip_varovalke = meritev.najdi_tip_varovalke()
                    t_varovalke = float(
                        meritev.najdi_t_varovalke().replace(" s", "").replace(",", ".")
                    )
                    i_varovalke = float(
                        meritev.najdi_I_varovalke().replace(" A", "").replace(",", ".")
                    )

                    Un = int(meritev.najdi_Un().replace(" V", ""))
                    Ipsc = meritev.najdi_Ia_Ipsc()
                    Z = meritev.najdi_Z()

                    ##

                    excel_delovna_datoteka = load_workbook(
                        "Meje za meritve.xlsx", data_only=True
                    )

                    t_varovalke_je_ustrezen = False
                    for vrednost in [0.1, 0.2, 0.4, 5.0]:
                        if t_varovalke == vrednost:
                            t_varovalke_je_ustrezen = True
                            break

                    if not t_varovalke_je_ustrezen:
                        tok_zascite = "X"
                        izracun = "X"

                    if tip_varovalke in ["gG", "gL"]:
                        excel_delovni_list = excel_delovna_datoteka["gG"]
                        prva_vrstica = 6
                        zadnja_vrstica = 34

                        slovar_t_varovalk_in_stolpcev = {
                            0.1: 1,
                            0.2: 4,
                            0.4: 7,
                            5.0: 10,
                        }
                        stolpec = slovar_t_varovalk_in_stolpcev[t_varovalke]

                        if t_varovalke == 0.1:
                            zadnja_vrstica = 30

                        stolpec_1 = [
                            excel_delovni_list.cell(row=i, column=stolpec).value
                            for i in range(prva_vrstica, zadnja_vrstica + 1)
                        ]

                        if i_varovalke not in stolpec_1:
                            tok_zascite = "X"
                            izracun = "X"
                        else:
                            tok_zascite = excel_delovni_list.cell(
                                row=stolpec_1.index(i_varovalke) + 6, column=stolpec + 1
                            ).value
                            izracun = 2 / 3 * (Un / tok_zascite)

                    if tip_varovalke in ["B", "D", "C"]:
                        excel_delovni_list = excel_delovna_datoteka["BCD"]

                        prva_vrstica = 6
                        zadnja_vrstica = 17

                        slovar_tipov_varovalk_in_stolpcev = {"B": 2, "C": 4, "D": 6}
                        stolpec = slovar_tipov_varovalk_in_stolpcev[tip_varovalke]

                        stolpec_1 = [
                            excel_delovni_list.cell(row=i, column=1).value
                            for i in range(prva_vrstica, zadnja_vrstica + 1)
                        ]

                        if i_varovalke not in stolpec_1:
                            tok_zascite = "X"
                            izracun = "X"
                        else:
                            vrednost_vrst = {"B": 1, "C": 3, "D": 5}
                            tok_zascite = excel_delovni_list.cell(
                                row=stolpec_1.index(i_varovalke) + 6,
                                column=stolpec + vrednost_vrst[tip_varovalke],
                            ).value
                            izracun = 2 / 3 * (Un / tok_zascite)

                    ##

                    if tip_varovalke in ["NV"]:
                        excel_delovni_list = excel_delovna_datoteka["NV"]

                        # TODO, ko dobiva naslednje tabele

                        prva_vrstica = 6
                        zadnja_vrstica = 17

                        slovar_tipov_varovalk_in_stolpcev = {"B": 2, "C": 4, "D": 6}
                        stolpec = slovar_tipov_varovalk_in_stolpcev[tip_varovalke]

                        stolpec_1 = [
                            excel_delovni_list.cell(row=i, column=1).value
                            for i in range(prva_vrstica, zadnja_vrstica + 1)
                        ]

                        if i_varovalke not in stolpec_1:
                            tok_zascite = "X"
                            izracun = "X"
                        else:
                            vrednost_vrst = {"B": 1, "C": 3, "D": 5}
                            tok_zascite = excel_delovni_list.cell(
                                row=stolpec_1.index(i_varovalke) + 6,
                                column=stolpec + vrednost_vrst[tip_varovalke],
                            ).value
                            izracun = 2 / 3 * (Un / tok_zascite)

                    array_ki_ga_zapisemo_v_csv = [
                        komentar,
                        PRAZNO,
                        Un,
                        tok_zascite,
                        izracun,
                        f"{Ipsc}/{Z}",
                        pot,
                    ]
                    writer.writerow(array_ki_ga_zapisemo_v_csv)
            csvfile.close()

    if slovar_vrst_meritev["R iso"] + slovar_vrst_meritev["R IZO"] > 0:
        with open(CSVFILE_RISO, "a", encoding="utf-8", newline="") as csvfile:
            writer = csv.writer(
                csvfile, delimiter=";", quotechar='"', quoting=csv.QUOTE_MINIMAL
            )

            for meritev in kocka:
                if meritev.doloci_vrsto_meritve() in ["R iso", "R IZO"]:
                    komentar = meritev.najdi_komentar()
                    riso = meritev.najdi_Riso()

                    array_ki_ga_zapisemo_v_csv = [PRAZNO, PRAZNO, riso, komentar]
                    writer.writerow(array_ki_ga_zapisemo_v_csv)
            csvfile.close()

    if (
        slovar_vrst_meritev["Discharge time"] + slovar_vrst_meritev["Čas praznjenja"]
        > 0
    ):
        with open(CSVFILE_DISCHARGE_TIME, "a", encoding="utf-8", newline="") as csvfile:
            writer = csv.writer(
                csvfile, delimiter=";", quotechar='"', quoting=csv.QUOTE_MINIMAL
            )

            for meritev in kocka:
                if meritev.doloci_vrsto_meritve() in [
                    "Discharge time",
                    "Čas praznjenja",
                ]:
                    komentar = meritev.najdi_komentar()

                    t = float(meritev.najdi_t().replace(",", "."))
                    meja_t = meritev.najdi_meja_t()

                    if meja_t == "5":
                        drugi_tretji = [t, PRAZNO]
                    elif meja_t == "1":
                        drugi_tretji = [PRAZNO, t]
                    else:
                        drugi_tretji = ["X", "X"]
                        print("meja_t ni niti 1s ali 5s")

                    array_ki_ga_zapisemo_v_csv = [PRAZNO, *drugi_tretji, komentar]
                    writer.writerow(array_ki_ga_zapisemo_v_csv)
            csvfile.close()


def najdi_po_vrsti_urejen_seznam_datumov(vse_besedilo):
    loceno_besedilo_po_presledkih = vse_besedilo.split()
    seznam_stringov_z_datumi = []
    for i in loceno_besedilo_po_presledkih:
        if re.search(r"\d{2}\.\d{2}\.\d{4}", i):
            # vsi datumi so pravilne oblike (zraven so zapisane odvečne ničle),
            # zato je upravičeno tole
            string_datuma = i[9:]
            # print(string_datuma)
            datum = datetime.strptime(string_datuma, "%d.%m.%Y")
            seznam_stringov_z_datumi.append(datum)
    return [
        datetime.strftime(datum, "%d.%m.%Y")
        for datum in sorted(seznam_stringov_z_datumi)
    ]


def najdi_pot_izven_razreda_Meritev(besedilo):
    idx = besedilo.find("Pot:")
    string_ki_ga_obdelujemo = besedilo[idx:]
    if "Page" in string_ki_ga_obdelujemo:
        idx = string_ki_ga_obdelujemo.find("Page")
        # poanta je v temu, da so spredaj odvečne črke
        string_ki_ga_obdelujemo = string_ki_ga_obdelujemo[:idx]
    if "Serijsko" in string_ki_ga_obdelujemo:
        idx = string_ki_ga_obdelujemo.find("Serijsko")
        string_ki_ga_obdelujemo = string_ki_ga_obdelujemo[:idx]
    return string_ki_ga_obdelujemo
