from openpyxl import load_workbook

RDECA = "FF0000"
ORANZNA = "FFA500"
MODRA = "99FFFF"


def preberi_float_z_mesta_v_seznamu(seznam, mesto, pomozno_mesto=None):
    if pomozno_mesto is not None:
        return (
            float(seznam[mesto].split("/")[pomozno_mesto].replace(",", "."))
            if seznam[mesto].split("/")[pomozno_mesto] != "X"
            else "X"
        )
    else:
        return float(seznam[mesto].replace(",", ".")) if seznam[mesto] != "X" else "X"


def preveri_meje_osnovne(seznam, trafo=True):
    # funkcija prejme seznam s 25 elementi, ter binarno vrendost trafo

    slovar_problematicnih_meritev = dict()

    # faktorja, ki določata občutljivost oranžne barve
    faktor_za_oranzno_barvo_tok = 1.5
    # 1.5 pomeni, da je barva oranžna v primeru,
    # ko je rezultat manjši od 1.5-kratnika najnižje dovoljene meje
    faktor_za_oranzno_barvo_upornost = 0.7
    # 0.7 pomeni, da je barva oranžna v primeru,
    # ko je rezultat večji od 0.7-kratnika najvišje dovoljene meje

    tip_varovalke = seznam[8]
    i_varovalke = preberi_float_z_mesta_v_seznamu(seznam, 9)
    t_varovalke = preberi_float_z_mesta_v_seznamu(seznam, 10)

    pot = seznam[25]

    # Če je luč, obravnavamo mejo drugače
    if "lamp" in pot or "Lamp" in pot:
        meja_du = 0.05 if trafo else 0.03
    else:
        meja_du = 0.07 if trafo else 0.05

    # Tukaj preimenuj spremenljivke, da bodo bolj opisne

    zs = preberi_float_z_mesta_v_seznamu(seznam, 11, pomozno_mesto=0)
    ik1 = preberi_float_z_mesta_v_seznamu(seznam, 11, pomozno_mesto=1)
    zl = preberi_float_z_mesta_v_seznamu(seznam, 12, pomozno_mesto=0)
    ik2 = preberi_float_z_mesta_v_seznamu(seznam, 12, pomozno_mesto=1)
    du = preberi_float_z_mesta_v_seznamu(seznam, 12, pomozno_mesto=2)
    if type(du) == float:
        du /= 100

    rlow = seznam[5]
    riso = seznam[7]

    if rlow != "X":
        if ">" in rlow:
            slovar_problematicnih_meritev[5] = RDECA
        else:
            rlow = float(rlow.replace(",", "."))
            if rlow > 1:
                slovar_problematicnih_meritev[5] = RDECA
            elif rlow > 0.8:
                slovar_problematicnih_meritev[5] = ORANZNA

    elif riso != "X":
        if ">" not in riso:
            riso = float(riso.replace(",", "."))
            if riso < 1:
                slovar_problematicnih_meritev[7] = RDECA
            elif riso < 2:
                slovar_problematicnih_meritev[7] = ORANZNA

    excel_delovna_datoteka = load_workbook("Meje za meritve.xlsx", data_only=True)

    # Označimo z rdečo neobstoječe čase varovalk

    t_varovalke_je_ustrezen = False
    for vrednost in [0.1, 0.2, 0.4, 5.0]:
        if t_varovalke == vrednost:
            t_varovalke_je_ustrezen = True
            break

    if not t_varovalke_je_ustrezen:
        slovar_problematicnih_meritev[10] = RDECA
        stolpec = 1

    if tip_varovalke in ["gG", "NV", "gL"]:
        excel_delovni_list = excel_delovna_datoteka["gG"]
        prva_vrstica = 6
        zadnja_vrstica = 34

        slovar_t_varovalk_in_stolpcev = {0.1: 1, 0.2: 4, 0.4: 7, 5.0: 10}
        stolpec = slovar_t_varovalk_in_stolpcev[t_varovalke]

        if t_varovalke == 0.1:
            zadnja_vrstica = 30

        stolpec_1 = [
            excel_delovni_list.cell(row=i, column=stolpec).value
            for i in range(prva_vrstica, zadnja_vrstica + 1)
        ]
        stolpec_2 = [
            excel_delovni_list.cell(row=i, column=stolpec + 1).value
            for i in range(prva_vrstica, zadnja_vrstica + 1)
        ]
        stolpec_3 = [
            excel_delovni_list.cell(row=i, column=stolpec + 2).value
            for i in range(prva_vrstica, zadnja_vrstica + 1)
        ]

        if i_varovalke not in stolpec_1:
            indeks_ujemajoce_varovalke = 0
            slovar_problematicnih_meritev[9] = RDECA
        else:
            indeks_ujemajoce_varovalke = stolpec_1.index(i_varovalke)

    if tip_varovalke in ["B", "D", "C"]:
        excel_delovni_list = excel_delovna_datoteka["BCD"]

        prva_vrstica = 6
        zadnja_vrstica = 17

        slovar_tipov_varovalk_in_stolpcev = {"B": 2, "C": 4, "D": 6}
        stolpec = slovar_tipov_varovalk_in_stolpcev[tip_varovalke]

        stolpec_1 = [
            excel_delovni_list.cell(row=i, column=1).value
            for i in range(prva_vrstica, zadnja_vrstica + 1)
        ]
        stolpec_2 = [
            excel_delovni_list.cell(row=i, column=stolpec).value
            for i in range(prva_vrstica, zadnja_vrstica + 1)
        ]
        stolpec_3 = [
            excel_delovni_list.cell(row=i, column=stolpec + 1).value
            for i in range(prva_vrstica, zadnja_vrstica + 1)
        ]

        if i_varovalke not in stolpec_1:
            indeks_ujemajoce_varovalke = 0
            slovar_problematicnih_meritev[9] = RDECA
        else:
            indeks_ujemajoce_varovalke = stolpec_1.index(i_varovalke)

    if ik1 == "X":
        slovar_problematicnih_meritev[11] = MODRA
    elif ik2 == "X":
        slovar_problematicnih_meritev[12] = MODRA
    elif t_varovalke_je_ustrezen:
        if ik1 < stolpec_2[indeks_ujemajoce_varovalke]:
            slovar_problematicnih_meritev[11] = RDECA
        elif ik1 <= stolpec_2[indeks_ujemajoce_varovalke] * faktor_za_oranzno_barvo_tok:
            slovar_problematicnih_meritev[11] = ORANZNA

        if ik2 < stolpec_2[indeks_ujemajoce_varovalke]:
            slovar_problematicnih_meritev[12] = RDECA
        elif ik2 == stolpec_2[indeks_ujemajoce_varovalke] * faktor_za_oranzno_barvo_tok:
            slovar_problematicnih_meritev[12] = ORANZNA

    if zs == "X":
        slovar_problematicnih_meritev[11] = MODRA
    elif zl == "X":
        slovar_problematicnih_meritev[12] = MODRA
    elif du == "X":
        slovar_problematicnih_meritev[12] = MODRA
    elif t_varovalke_je_ustrezen:
        if zl > stolpec_3[indeks_ujemajoce_varovalke]:
            slovar_problematicnih_meritev[12] = RDECA
        elif (
            zl
            >= stolpec_3[indeks_ujemajoce_varovalke] * faktor_za_oranzno_barvo_upornost
        ):
            if slovar_problematicnih_meritev.get(12, "") != RDECA:
                slovar_problematicnih_meritev[12] = ORANZNA
        if zs > stolpec_3[indeks_ujemajoce_varovalke]:
            slovar_problematicnih_meritev[11] = RDECA
        elif (
            zs
            >= stolpec_3[indeks_ujemajoce_varovalke] * faktor_za_oranzno_barvo_upornost
        ):
            if slovar_problematicnih_meritev.get(11, "") != RDECA:
                slovar_problematicnih_meritev[11] = ORANZNA

        if du > meja_du:
            slovar_problematicnih_meritev[12] = RDECA
        elif du >= meja_du * faktor_za_oranzno_barvo_upornost:
            if slovar_problematicnih_meritev.get(12, "") != RDECA:
                slovar_problematicnih_meritev[12] = ORANZNA

    return slovar_problematicnih_meritev


def preveri_meje_RLOW4(seznam, trafo=True):

    slovar_problematicnih_meritev = dict()
    r = seznam[1].replace(",", ".")

    if r != "":
        if ">" in r:
            slovar_problematicnih_meritev[1] = RDECA
        else:
            r = float(r) if r != "X" else "X"

            if r == "X":
                slovar_problematicnih_meritev[1] = MODRA
            elif r > 1:
                slovar_problematicnih_meritev[1] = RDECA
            elif r > 0.8:
                slovar_problematicnih_meritev[1] = ORANZNA

    return slovar_problematicnih_meritev
