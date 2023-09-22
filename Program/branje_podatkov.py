# datoteka za branje podatkov z datoteke in generiranje ustrezne csv datoteke
from model import *
import Meritev
from colorama import Fore
from meje_instalacije import *
from meje_stroji import *
from meje_strelovodi import *
from meje_omare import *
import re
import os
from poti_do_datotek import *

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

# Ko se importira GUI, se avtomatično izvede vprašalnik
from GUI import *

# Če hočeš vklopiti debugganje, daš debug_mode na True
debug_mode = True

# obstaja 7 osnovnih vrst meritev: AUTO TN, Zloop mΩ, Z LINE, RCD Auto, R low 4, Varistor, R iso, R IZO, Z loop 4W, Z line 4W

if "trafo_postaja" in locals() and trafo_postaja:
    print(Fore.GREEN + "--------------------------------------")
    print("|| Objekt ima svojo trafo postajo. || ")
    print("--------------------------------------" + Fore.WHITE)

else:
    print(Fore.RED + "--------------------------------------")
    print("|| Objekt nima svoje trafo postaje. || ")
    print("--------------------------------------" + Fore.WHITE)

with open("Podatki_z_merjenj.txt", encoding="utf-8") as podatki:
    celotno_besedilo_z_merjenj = (
        podatki.read()
        .replace(", (+)", ",(+)")
        .replace(", (-)", ",(-)")
        .replace(";", "-")
        # .replace("Z loop 4W", "Zloop")
        # .replace("Z line 4W", "Z LINE")
        .replace("\n", " ")
        .replace("Pot:", "🙈Pot:")
    )

    # prvi string v seznamu e vedno prazen, zato geldaš samo naprej
    loceno_besedilo_na_kocke_teksta = celotno_besedilo_z_merjenj.split("🙈")[1:]

    seznam_datumov_po_vrstnem_redu = najdi_po_vrsti_urejen_seznam_datumov(
        celotno_besedilo_z_merjenj
    )

# tukaj se da še polepšati zapis datuma
print(
    Fore.YELLOW + "Meritve so bile opravljene od:",
    seznam_datumov_po_vrstnem_redu[0],
    "do:",
    seznam_datumov_po_vrstnem_redu[-1],
)

seznam_vseh_kock_meritev_brez_poti_na_koncu = []
seznam_ustreznih_poti_do_kock_meritev = []

# Tole je treba prenesti v model!


def ustvari_seznam_vseh_meritev():
    seznam_vseh_meritev = []

    for kocka_teksta in loceno_besedilo_na_kocke_teksta:
        slovar_meritev = {i: 0 for i in SEZNAM_VRST_MERITEV}
        pot_do_druzine_meritev = najdi_pot_izven_razreda_Meritev(kocka_teksta)
        print(pot_do_druzine_meritev)

        for vrsta_meritve in SEZNAM_VRST_MERITEV:
            if vrsta_meritve in kocka_teksta:
                slovar_meritev[vrsta_meritve] = kocka_teksta.count(vrsta_meritve)

        stevilo_meritev_v_trenutni_kocki = sum(slovar_meritev.values())
        print(stevilo_meritev_v_trenutni_kocki)
        if stevilo_meritev_v_trenutni_kocki == 1:
            # Ločimo primere glede na število meritev v kocki
            if kocka_teksta.count("p//") == 0:
                seznam_vseh_meritev.append(
                    [
                        Meritev.Meritev(
                            # Tukaj je skoraj sigurno nekaj preveč v replaceu
                            kocka_teksta.replace("\n", " ")
                            .replace("\r\n", " ")
                            .strip()
                        )
                    ]
                )
                seznam_ustreznih_poti_do_kock_meritev.append(pot_do_druzine_meritev)

        else:
            seznam_indeksov_posameznih_meritev = []
            seznam_vseh_kock_meritev_brez_poti_na_koncu = []
            for key in slovar_meritev:
                seznam_indeksov_posameznih_meritev += [
                    m.start() for m in re.finditer(key, kocka_teksta)
                ]
            seznam_indeksov_posameznih_meritev.sort()
            seznam_vseh_kock_meritev_brez_poti_na_koncu += [
                kocka_teksta[i:j]
                for i, j in zip(
                    seznam_indeksov_posameznih_meritev,
                    seznam_indeksov_posameznih_meritev[1:] + [None],
                )
            ]

            zacasen_seznam_za_vse_meritve = []
            for meritev in seznam_vseh_kock_meritev_brez_poti_na_koncu:
                if meritev.count("p//") == 0:
                    zacasen_seznam_za_vse_meritve.append(
                        Meritev.Meritev(
                            meritev.replace("\n", " ").replace("\r\n", " ").strip()
                        )
                    )
                # To zna biti malenkost nevarno, ampak se zdi da deluje
                elif meritev.count("p//") > 0 and meritev.count("Z LINE") > 0:
                    zacasen_seznam_za_vse_meritve.append(
                        Meritev.Meritev(
                            meritev.replace("\n", " ").replace("\r\n", " ").strip()
                        )
                    )

            if zacasen_seznam_za_vse_meritve:
                seznam_vseh_meritev.append(zacasen_seznam_za_vse_meritve)
                seznam_ustreznih_poti_do_kock_meritev.append(pot_do_druzine_meritev)

    return seznam_vseh_meritev


# seznam_vseh_meritev so vse meritve,
# vsaki kocki lahko priprada 1 ali več meritev
seznam_vseh_meritev = ustvari_seznam_vseh_meritev()
slovar_kock_in_ustreznih_poti = dict(
    zip(range(len(seznam_vseh_meritev)), seznam_ustreznih_poti_do_kock_meritev)
)
print(
    "Število vseh kock:",
    len(seznam_vseh_meritev),
    "\nŠtevilo ustreznih poti do kock:",
    len(seznam_ustreznih_poti_do_kock_meritev),
    "" + Fore.WHITE,
)


match vrsta_stroja:
    case "električna omara":
        with open(
            CSVFILE_ELEKTRICNE_OMARE,
            "w",
            encoding="utf-8",
            newline="",
        ) as csvfile:
            csvfile.close()

        excel_delovna_datoteka = load_workbook(TEMPLATE_ELEKTRICNE_OMARE)
        excel_delovna_datoteka.save(EXCELFILE_ELEKTRICNE_OMARE)
        excel_delovna_datoteka.close()

        for kocka in seznam_vseh_meritev:
            zapisi_kocko_meritev_v_excel_elektricne_omare(
                kocka,
                seznam_vseh_meritev,
                slovar_kock_in_ustreznih_poti,
                prevedi_v_anglescino,
            )

        with open(
            CSVFILE_ELEKTRICNE_OMARE,
            "r",
            encoding="utf-8",
            newline="",
        ) as csvfile:
            vrstice = csvfile.readlines()

        excel_delovna_datoteka = load_workbook(EXCELFILE_ELEKTRICNE_OMARE)

        excel_delovni_list = excel_delovna_datoteka.active
        dolzina_templata = excel_delovni_list.max_column
        visina_templata = excel_delovni_list.max_row

        sekcija = ""
        seznam_sekcij = []
        st_vrstic = 0

        for i, vrstica in enumerate(vrstice):
            pot_locena_na_elemente = vrstica.replace("\n", " ").strip().split("//")
            for element in pot_locena_na_elemente:
                if "Sekcija: " in element.strip():
                    sekcija = element[
                        element.index("Sekcija:") + len("Sekcija:") :
                    ].replace("Sekcija: ", "")

                if (
                    seznam_sekcij == []
                    or sekcija != ""
                    and sekcija != seznam_sekcij[-1]
                ):
                    seznam_sekcij.append(sekcija)
                    excel_delovni_list.append([f"{sekcija}"])
                    excel_delovni_list.merge_cells(
                        start_row=i + visina_templata + len(seznam_sekcij),
                        start_column=1,
                        end_row=i + visina_templata + len(seznam_sekcij),
                        end_column=dolzina_templata,
                    )
                    top_cell = excel_delovni_list[
                        f"A{i  + visina_templata + len(seznam_sekcij)}"
                    ]
                    top_cell.alignment = Alignment(
                        horizontal="center", vertical="center"
                    )
                    top_cell.font = Font(b=True)
                    top_cell.fill = PatternFill(
                        start_color=RUMENO_BEZ, end_color=RUMENO_BEZ, fill_type="solid"
                    )
            excel_delovni_list.append(vrstica.split(";"))
            st_vrstic += 1

            napake = ""
            str_napake = f'napake = preveri_meje_omare(vrstica.split(";"))'
            if debug_mode:  # Zelo uporabno pri debugganju.
                print(pot_locena_na_elemente)
            exec(str_napake)
            if napake:
                for indeks, barva in napake.items():
                    excel_delovni_list.cell(
                        i + visina_templata + len(seznam_sekcij) + 1, indeks + 1
                    ).fill = PatternFill(
                        start_color=barva, end_color=barva, fill_type="solid"
                    )

        excel_delovna_datoteka.save(EXCELFILE_ELEKTRICNE_OMARE)

        print(f"Število vseh meritev: {st_vrstic}")

    # ---------------------------------------------------------------------------------------------------------------------------
    # ---------------------------------------------------------------------------------------------------------------------------
    # ---------------------------------------------------------------------------------------------------------------------------

    case "inštalacija":
        VSE_PRIPONE_DATOTEK = ["osnovne", "RCD", "RLOW4", "VARISTOR"]
        for pripona in VSE_PRIPONE_DATOTEK:
            with open(
                os.path.join(
                    "..", "Csvji", "Instalacije", f"csv_za_excel_datoteko_{pripona}.csv"
                ),
                "w",
                encoding="utf-8",
                newline="",
            ) as csvfile:
                csvfile.close()

            excel_delovna_datoteka = load_workbook(
                os.path.join(
                    "..",
                    "Templati",
                    "Instalacije",
                    f"template_za_{pripona}_meritve.xlsx",
                )
            )
            excel_delovna_datoteka.save(
                os.path.join(
                    "..", "Porocila", "Instalacije", f"excel_datoteka_{pripona}.xlsx"
                )
            )
            excel_delovna_datoteka.close()

        for kocka in seznam_vseh_meritev:
            zapisi_kocko_meritev_v_excel_instalacije(
                kocka, seznam_vseh_meritev, slovar_kock_in_ustreznih_poti
            )

        for pripona in VSE_PRIPONE_DATOTEK:
            with open(
                os.path.join(
                    "..", "Csvji", "Instalacije", f"csv_za_excel_datoteko_{pripona}.csv"
                ),
                "r",
                encoding="utf-8",
                newline="",
            ) as csvfile:
                vrstice = csvfile.readlines()

            excel_delovna_datoteka = load_workbook(
                os.path.join(
                    "..", "Porocila", "Instalacije", f"excel_datoteka_{pripona}.xlsx"
                )
            )
            excel_delovni_list = excel_delovna_datoteka.active
            dolzina_templata = excel_delovni_list.max_column
            visina_templata = excel_delovni_list.max_row

            sekcija = ""
            seznam_sekcij = []

            for i, vrstica in enumerate(vrstice):
                pot_locena_na_elemente = vrstica.replace("\n", " ").strip().split("//")
                for element in pot_locena_na_elemente:
                    if "Sekcija: " in element.strip():
                        sekcija = element[
                            element.index("Sekcija:") + len("Sekcija:") :
                        ].replace("Sekcija: ", "")

                    if (
                        seznam_sekcij == []
                        or sekcija != ""
                        and sekcija != seznam_sekcij[-1]
                    ):
                        seznam_sekcij.append(sekcija)
                        excel_delovni_list.append([f"{sekcija}"])
                        excel_delovni_list.merge_cells(
                            start_row=i + visina_templata + len(seznam_sekcij),
                            start_column=1,
                            end_row=i + visina_templata + len(seznam_sekcij),
                            end_column=dolzina_templata,
                        )
                        top_cell = excel_delovni_list[
                            f"A{i  + visina_templata + len(seznam_sekcij)}"
                        ]
                        top_cell.alignment = Alignment(
                            horizontal="center", vertical="center"
                        )
                        top_cell.font = Font(b=True)
                        top_cell.fill = PatternFill(
                            start_color=RUMENO_BEZ,
                            end_color=RUMENO_BEZ,
                            fill_type="solid",
                        )
                excel_delovni_list.append(vrstica.split(";"))

                # avtomatično barvanje je samo v primeru, ko ni VARISTOR
                if pripona not in ["RCD", "VARISTOR"]:
                    napake = ""
                    str_napake = f'napake = preveri_meje_{pripona}(vrstica.split(";"), trafo=trafo_postaja)'
                    if debug_mode:  # Zelo uporabno pri debugganju.
                        print(pot_locena_na_elemente)
                    exec(str_napake)
                    if napake:
                        for indeks, barva in napake.items():
                            excel_delovni_list.cell(
                                i + visina_templata + len(seznam_sekcij) + 1, indeks + 1
                            ).fill = PatternFill(
                                start_color=barva, end_color=barva, fill_type="solid"
                            )

            excel_delovna_datoteka.save(
                os.path.join(
                    "..", "Porocila", "Instalacije", f"excel_datoteka_{pripona}.xlsx"
                )
            )

    # ---------------------------------------------------------------------------------------------------------------------------
    # ---------------------------------------------------------------------------------------------------------------------------
    # ---------------------------------------------------------------------------------------------------------------------------

    case "stroj":
        VSE_PRIPONE_DATOTEK = ["ZLOOP", "R ISO", "DISCHARGE TIME", "NEPREKINJENOST"]
        for pripona in VSE_PRIPONE_DATOTEK:
            with open(
                os.path.join(
                    "..", "Csvji", "Stroji", f"csv_za_excel_datoteko_{pripona}.csv"
                ),
                "w",
                encoding="utf-8",
                newline="",
            ) as csvfile:
                csvfile.close()

            excel_delovna_datoteka = load_workbook(
                os.path.join(
                    "..", "Templati", "Stroji", f"template_za_{pripona}_meritve.xlsx"
                )
            )
            excel_delovna_datoteka.save(
                os.path.join(
                    "..", "Porocila", "Stroji", f"excel_datoteka_{pripona}.xlsx"
                )
            )
            excel_delovna_datoteka.close()

        for kocka in seznam_vseh_meritev:
            zapisi_kocko_meritev_v_excel_stroji(
                kocka,
                seznam_vseh_meritev,
                slovar_kock_in_ustreznih_poti,
                t_varovalke_neprekinjenost=t_varovalke_stroji_neprekinjenost,
                I_varovalke_neprekinjenost=I_varovalke_stroji_neprekinjenost,
                tip_varovalke_neprekinjenost=tip_varovalke_stroji_neprekinjenost,
                meja_izolacijske_upornosti_stroji_riso_rdeca=meja_izolacijske_upornosti_stroji_riso_rdeca,
                prevedi_v_anglescino=prevedi_v_anglescino,
                napetost_dotika=napetost_dotika,
            )

        for pripona in VSE_PRIPONE_DATOTEK:
            with open(
                os.path.join(
                    "..", "Csvji", "Stroji", f"csv_za_excel_datoteko_{pripona}.csv"
                ),
                "r",
                encoding="utf-8",
                newline="",
            ) as csvfile:
                vrstice = csvfile.readlines()

            excel_delovna_datoteka = load_workbook(
                os.path.join(
                    "..", "Porocila", "Stroji", f"excel_datoteka_{pripona}.xlsx"
                )
            )

            excel_delovni_list = excel_delovna_datoteka.active
            dolzina_templata = excel_delovni_list.max_column
            visina_templata = excel_delovni_list.max_row

            sekcija = ""
            seznam_sekcij = []

            for i, vrstica in enumerate(vrstice):
                pot_locena_na_elemente = vrstica.replace("\n", " ").strip().split("//")
                for element in pot_locena_na_elemente:
                    if "Sekcija: " in element.strip():
                        sekcija = element[
                            element.index("Sekcija:") + len("Sekcija:") :
                        ].replace("Sekcija: ", "")

                    if (
                        seznam_sekcij == []
                        or sekcija != ""
                        and sekcija != seznam_sekcij[-1]
                    ):
                        seznam_sekcij.append(sekcija)
                        excel_delovni_list.append([f"{sekcija}"])
                        excel_delovni_list.merge_cells(
                            start_row=i + visina_templata + len(seznam_sekcij),
                            start_column=1,
                            end_row=i + visina_templata + len(seznam_sekcij),
                            end_column=dolzina_templata,
                        )
                        top_cell = excel_delovni_list[
                            f"A{i  + visina_templata + len(seznam_sekcij)}"
                        ]
                        top_cell.alignment = Alignment(
                            horizontal="center", vertical="center"
                        )
                        top_cell.font = Font(b=True)
                        top_cell.fill = PatternFill(
                            start_color=RUMENO_BEZ,
                            end_color=RUMENO_BEZ,
                            fill_type="solid",
                        )
                excel_delovni_list.append(vrstica.split(";"))

                # DISCHARGE TIME na roko
                if pripona in ["R ISO", "ZLOOP", "NEPREKINJENOST"]:
                    napake = ""
                    if debug_mode:
                        print(pot_locena_na_elemente)
                    if pripona == "R ISO":
                        str_napake = f'napake = preveri_meje_R_ISO(vrstica.split(";"), meja_izolacijske_upornosti_stroji_riso_oranzna = meja_izolacijske_upornosti_stroji_riso_oranzna)'
                    else:
                        str_napake = (
                            f'napake = preveri_meje_{pripona}(vrstica.split(";"))'
                        )
                    exec(str_napake)
                    if napake:
                        for indeks, barva in napake.items():
                            excel_delovni_list.cell(
                                i + visina_templata + len(seznam_sekcij) + 1, indeks + 1
                            ).fill = PatternFill(
                                start_color=barva, end_color=barva, fill_type="solid"
                            )

                excel_delovna_datoteka.save(
                    os.path.join(
                        "..", "Porocila", "Stroji", f"excel_datoteka_{pripona}.xlsx"
                    )
                )

    # ---------------------------------------------------------------------------------------------------------------------------
    # ---------------------------------------------------------------------------------------------------------------------------
    # ---------------------------------------------------------------------------------------------------------------------------
    case "strelovod":
        with open(
            CSVFILE_STRELOVODI,
            "w",
            encoding="utf-8",
            newline="",
        ) as csvfile:
            csvfile.close()

        excel_delovna_datoteka = load_workbook(TEMPLATE_STRELOVODI)
        excel_delovna_datoteka.save(EXCELFILE_STRELOVODI)
        excel_delovna_datoteka.close()

        for kocka in seznam_vseh_meritev:
            zapisi_kocko_meritev_v_excel_strelovodi(
                kocka, seznam_vseh_meritev, slovar_kock_in_ustreznih_poti
            )

        with open(
            CSVFILE_STRELOVODI,
            "r",
            encoding="utf-8",
            newline="",
        ) as csvfile:
            vrstice = csvfile.readlines()

        excel_delovna_datoteka = load_workbook(EXCELFILE_STRELOVODI)

        excel_delovni_list = excel_delovna_datoteka.active
        dolzina_templata = excel_delovni_list.max_column
        visina_templata = excel_delovni_list.max_row

        sekcija = ""
        seznam_sekcij = []

        for i, vrstica in enumerate(vrstice):
            pot_locena_na_elemente = vrstica.replace("\n", " ").strip().split("//")
            for element in pot_locena_na_elemente:
                if "Sekcija: " in element.strip():
                    sekcija = element[
                        element.index("Sekcija:") + len("Sekcija:") :
                    ].replace("Sekcija: ", "")

                if (
                    seznam_sekcij == []
                    or sekcija != ""
                    and sekcija != seznam_sekcij[-1]
                ):
                    seznam_sekcij.append(sekcija)
                    excel_delovni_list.append([f"{sekcija}"])
                    excel_delovni_list.merge_cells(
                        start_row=i + visina_templata + len(seznam_sekcij),
                        start_column=1,
                        end_row=i + visina_templata + len(seznam_sekcij),
                        end_column=dolzina_templata,
                    )
                    top_cell = excel_delovni_list[
                        f"A{i  + visina_templata + len(seznam_sekcij)}"
                    ]
                    top_cell.alignment = Alignment(
                        horizontal="center", vertical="center"
                    )
                    top_cell.font = Font(b=True)
                    top_cell.fill = PatternFill(
                        start_color=RUMENO_BEZ, end_color=RUMENO_BEZ, fill_type="solid"
                    )
            excel_delovni_list.append(vrstica.split(";"))

            napake = ""
            str_napake = f'napake = preveri_meje_strelovodi(vrstica.split(";"))'
            if debug_mode:  # Zelo uporabno pri debugganju.
                print(pot_locena_na_elemente)
            exec(str_napake)
            if napake:
                for indeks, barva in napake.items():
                    excel_delovni_list.cell(
                        i + visina_templata + len(seznam_sekcij) + 1, indeks + 1
                    ).fill = PatternFill(
                        start_color=barva, end_color=barva, fill_type="solid"
                    )

        excel_delovna_datoteka.save(EXCELFILE_STRELOVODI)
    case _:
        # To se itak ne more zgoditi
        print("Nekaj moraš izbrati!")
