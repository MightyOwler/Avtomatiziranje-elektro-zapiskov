import re
from datetime import datetime
import csv
from Meritev import pretvori_v_osnovne_enote, SEZNAM_VRST_MERITEV, Meritev
from poti_do_datotek import *
from openpyxl import load_workbook
import json

RUMENO_BEZ = "F5C77E"

MEJE_ZA_MERITVE = os.path.join("..", "Meje za meritve.xlsx")

# V slovarju so na levi strani slovenske besede, na desni pa angleške.

jsonfile = open("slovar_besed.json")
SLOVAR = json.load(jsonfile)


def prevedi_s_slovarjem(string, prevedi_v_anglescino):
    if prevedi_v_anglescino:
        for beseda in SLOVAR:
            string = string.replace(beseda, SLOVAR[beseda].lower())
            string = string.replace(beseda.upper(), SLOVAR[beseda].lower())
        return string.upper()
    else:
        return string.upper()


def oklesti_komentar(komentar, nezazeljeni_string):
    """
    Od neke pojavitve stringa dalje vse zbriše
    """
    if nezazeljeni_string in komentar:
        komentar = komentar[0 : komentar.index(nezazeljeni_string)]
    return komentar


PRAZNO = " "

st_vnesenih_meritev = 0
st_vnesenih_meritev_RCD = 0


# Poanta je, da mora ostati v milisekundah, ne sme se pretovoriti.
def pretvori_string_milisekund_v_ustrezen_format(string):
    return string.replace(",", ".").replace(" ms", "").replace(">", "")


def vrednoti_string(s):
    """
    Funkcija, ki nam omogoča da razvrstimo meritve po velikosti kljub znaku >
    (to je konkretno pomembno pri RISO in RLOW4)
    """

    if ">" in s:
        return 1000000000000
    if "X" in s:
        return 0
    else:
        return float(pretvori_v_osnovne_enote(s.replace(",", ".")).replace(",", "."))


def pridobi_potne_elemente(kocka, loceno_besedilo, slovar_kock_in_ustreznih_poti):
    """
    Vrne ustrezne potne elemente znotraj posameznih funkcij
    """
    komentar = ""
    ime = "X"
    pot = (
        slovar_kock_in_ustreznih_poti[loceno_besedilo.index(kocka)]
        .replace("\n", " ")
        .strip()
    )
    pot_locena_na_elemente = pot.replace("\n", " ").strip().split("//")
    for element in pot_locena_na_elemente:
        if "Imenovanje: " in element.strip():
            ime = element.replace("Imenovanje: ", "")
            if "Circuit F" in ime:
                ime = ime.replace("Circuit ", "")
            elif re.search(r"Circuit\d", ime):
                ime = ime.replace("Circuit", "F")

    vrste_meritev_v_kocki = [meritev.doloci_vrsto_meritve() for meritev in kocka]
    slovar_vrst_meritev = {
        i: vrste_meritev_v_kocki.count(i) for i in SEZNAM_VRST_MERITEV
    }
    return (
        komentar,
        pot,
        pot_locena_na_elemente,
        ime,
        vrste_meritev_v_kocki,
        slovar_vrst_meritev,
    )


def zapisi_kocko_meritev_v_excel_instalacije(
    kocka, loceno_besedilo, slovar_kock_in_ustreznih_poti
):
    """
    Zapiše meritev v csv datoteko, ki je primerna za obdelavo v csv-ju

    Args: kocka, loceno_besedilo, slovar_kock_in_ustreznih_poti
    """

    global st_vnesenih_meritev
    global st_vnesenih_meritev_RCD

    (
        uln,
        zln,
        ipsc_ln,
        ipsc_lpe,
        rlpe,
        dU,
        zlpe,
        glavna_izenac_povezava,
        I_dN,
        t1x,
        t5x,
        Uc,
        ia_psc_navidezni_stolpec,
        maxRplusRminus,
        tip_varovalke,
        I_varovalke,
        t_varovalke,
        isc_faktor,
    ) = ("X" for _ in range(18))

    # Treba je pridobiti pot, pot ločeno na elemente, ime ter vrste meritev v kocki in slovar vrst meritev

    (
        komentar,
        pot,
        pot_locena_na_elemente,
        ime,
        vrste_meritev_v_kocki,
        slovar_vrst_meritev,
    ) = pridobi_potne_elemente(kocka, loceno_besedilo, slovar_kock_in_ustreznih_poti)

    # Najprej pogleda za R iso

    if slovar_vrst_meritev["R iso"] + slovar_vrst_meritev["R IZO"] > 1:
        nova_kocka = []
        seznam_riso_meritev = []
        for meritev in kocka:
            if meritev.doloci_vrsto_meritve() in ["R iso", "R IZO"]:
                # inšturment vedno izpiše v MΩ
                seznam_riso_meritev.append(
                    meritev.najdi_Rlpe().replace(" MΩ", "").replace(",", ".")
                )
        seznam_riso_meritev.sort(key=lambda x: vrednoti_string(x))
        riso_meritev_z_minimalno = seznam_riso_meritev[0]
        for idx, meritev in enumerate(kocka):
            if meritev.doloci_vrsto_meritve() in ["R iso", "R IZO"]:
                if idx == seznam_riso_meritev.index(riso_meritev_z_minimalno):
                    nova_kocka.append(meritev)
            else:
                nova_kocka.append(meritev)

        kocka = nova_kocka

    if slovar_vrst_meritev["R low 4"] > 0:
        with open(
            CSVFILE_INSTALACIJE_RLOW4, "a", encoding="utf-8", newline=""
        ) as csvfile:
            writer = csv.writer(
                csvfile, delimiter=";", quotechar='"', quoting=csv.QUOTE_MINIMAL
            )
            seznam_Rjev_za_rlow4_meritve = []

            # Spodnji loop se da močno izboljšati, veliko stvari je nekoliko nenavadnih

            for meritev in kocka:
                if meritev.doloci_vrsto_meritve() == "R low 4":
                    komentar = meritev.najdi_komentar()
                    seznam_Rjev_za_rlow4_meritve.append(meritev.najdi_R())
                    if (
                        # >1999 je zgornja meja, nastavljena na napravi
                        ">1999" in meritev.najdi_R_pozitivno()
                        or ">1999" in meritev.najdi_R_negativno()
                    ):
                        maxRplusRminus = ">1999"
                    else:
                        R_pozitivno_int = (
                            meritev.najdi_R_pozitivno()
                            .replace(",", ".")
                            .replace(" Ω", "")
                            .replace(">", "")
                        )
                        if R_pozitivno_int != "X":
                            R_pozitivno_int = float(R_pozitivno_int)

                        R_negativno_int = (
                            meritev.najdi_R_negativno()
                            .replace(",", ".")
                            .replace(" Ω", "")
                            .replace(">", "")
                        )
                        if R_negativno_int != "X":
                            R_negativno_int = float(R_negativno_int)

                        # Pri kateri velikosti spremenimo v int?
                        if R_negativno_int == "X" or R_pozitivno_int == "X":
                            maxRplusRminus = "X"
                        else:
                            if max(R_negativno_int, R_pozitivno_int) >= 100:
                                maxRplusRminus = (
                                    f"{int(max(R_negativno_int, R_pozitivno_int))}"
                                )
                            else:
                                maxRplusRminus = (
                                    f"{max(R_negativno_int, R_pozitivno_int)}"
                                )
                    R = meritev.najdi_R()
                    vrsta_meritve = meritev.doloci_vrsto_meritve()
                    array_ki_ga_zapisemo_v_csv = [
                        ime,
                        R,
                        maxRplusRminus,
                        komentar,
                        pot,
                    ]
                    writer.writerow(array_ki_ga_zapisemo_v_csv)
            csvfile.close()
            seznam_Rjev_za_rlow4_meritve.sort(key=lambda x: vrednoti_string(x))
            rlow4_meritev_z_minimalno = seznam_Rjev_za_rlow4_meritve[0]

        if slovar_vrst_meritev["RCD Auto"] > 1:
            print("Napaka: Imamo 2 ali več RCD Auto meritvi v eni kocki!")

        # To je napisano na daljši način zaradi preglednosti.
        for idx, vrednost_meritve in enumerate(seznam_Rjev_za_rlow4_meritve):
            if idx == seznam_Rjev_za_rlow4_meritve.index(rlow4_meritev_z_minimalno):
                # Glavno izenačitveno povezavo potrebujemo samo pri eni meritvi
                glavna_izenac_povezava = vrednost_meritve
                break

    for meritev in kocka:
        komentar = meritev.najdi_komentar()
        vrsta_meritve = meritev.doloci_vrsto_meritve()
        if vrsta_meritve in ["R iso", "R IZO"]:
            if vrsta_meritve == "R iso":
                rlpe = meritev.najdi_Rlpe().replace(" MΩ", "")
            else:
                rlpe = meritev.najdi_Riso().replace(" MΩ", "")

        if vrsta_meritve == "Padec napetosti":
            dU = meritev.najdi_dU()

        if vrsta_meritve == "RCD Auto":
            Uc = meritev.najdi_Uc()
            I_dN = meritev.najdi_I_dN()
            t1x_plus = pretvori_string_milisekund_v_ustrezen_format(
                meritev.najdi_t_IΔN_x1_plus()
            )
            t1x_neg = pretvori_string_milisekund_v_ustrezen_format(
                meritev.najdi_t_IΔN_x1_minus()
            )
            t5x_plus = pretvori_string_milisekund_v_ustrezen_format(
                meritev.najdi_t_IΔN_x5_plus()
            )
            t5x_neg = pretvori_string_milisekund_v_ustrezen_format(
                meritev.najdi_t_IΔN_x5_minus()
            )

            # V primeru, da vse količine obstajajo
            if "X" not in [t1x_plus, t1x_neg, t5x_plus, t5x_neg]:
                t1x_plus = float(t1x_plus)
                t1x_neg = float(t1x_neg)
                t5x_plus = float(t5x_plus)
                t5x_neg = float(t5x_neg)
                max_t1 = max(t1x_plus, t1x_neg)
                max_t5 = max(t5x_plus, t5x_neg)

                # Pri vrednosti večji ali enaki 100 rezultat zaokrožimo
                if max_t1 >= 100:
                    t1x = f"{int(max_t1)}".replace(".", ",")
                else:
                    t1x = f"{max_t1}".replace(".", ",")

                if max_t5 >= 100:
                    t5x = f"{int(max_t5)}".replace(".", ",")
                else:
                    t5x = f"{max_t5}".replace(".", ",")
            else:
                t1x = "X"
                t5x = "X"

        if vrsta_meritve == "Varistor":
            with open(
                CSVFILE_INSTALACIJE_VARISTOR,
                "a",
                encoding="utf-8",
                newline="",
            ) as csvfile:
                writer = csv.writer(
                    csvfile, delimiter=";", quotechar='"', quoting=csv.QUOTE_MINIMAL
                )

                uac = meritev.najdi_Uac()
                udc = meritev.najdi_Udc()

                array_ki_ga_zapisemo_v_csv = [
                    ime,
                    uac,
                    udc,
                    komentar,
                    vrsta_meritve,
                    pot,
                ]
                writer.writerow(array_ki_ga_zapisemo_v_csv)
                csvfile.close()

    # Stvar je treba narediti v večih korakih, saj lahko podatki v posamezni meritvi vplivajo na celotno kocko ali kasnejše.
    # najprej odpravimo AUTO TN

    for meritev in kocka:
        vrsta_meritve = meritev.doloci_vrsto_meritve()

        if vrsta_meritve == "AUTO TN":
            with open(
                CSVFILE_INSTALACIJE_OSNOVNE,
                "a",
                encoding="utf-8",
                newline="",
            ) as csvfile:
                writer = csv.writer(
                    csvfile, delimiter=";", quotechar='"', quoting=csv.QUOTE_MINIMAL
                )

                # Tole je treba izboljšati
                glavna_izenac_povezava = meritev.najdi_R()
                uln = meritev.najdi_Uln()
                zln = meritev.najdi_Z_LN()
                ipsc_ln = meritev.najdi_Ipsc_LN()
                ipsc_lpe = meritev.najdi_Ipsc_LPE()
                dU = meritev.najdi_dU()
                zlpe = meritev.najdi_Z_LPE()
                ia_psc_navidezni_stolpec = meritev.najdi_Ia_Ipsc()
                tip_varovalke = meritev.najdi_tip_varovalke()
                I_varovalke = meritev.najdi_I_varovalke()
                t_varovalke = meritev.najdi_t_varovalke()
                isc_faktor = meritev.najdi_Isc_faktor()
                komentar = meritev.najdi_komentar()

                st_vnesenih_meritev += 1
                array_ki_ga_zapisemo_v_csv = [
                    st_vnesenih_meritev,
                    ime,
                    PRAZNO,
                    PRAZNO,
                    PRAZNO,
                    glavna_izenac_povezava,
                    PRAZNO,
                    rlpe,
                    tip_varovalke,
                    I_varovalke,
                    t_varovalke,
                    f"{zlpe}/{ipsc_ln}",
                    f"{zln}/{ipsc_lpe}/{dU}",
                    PRAZNO,
                    I_dN,
                    PRAZNO,
                    t1x,
                    t5x,
                    Uc,
                    PRAZNO,
                    komentar,
                    vrsta_meritve,
                    uln,
                    maxRplusRminus,
                    isc_faktor,
                    ia_psc_navidezni_stolpec,
                    pot,
                ]
                writer.writerow(array_ki_ga_zapisemo_v_csv)
                csvfile.close()

        # Sočasno odpravljamo tudi Z auto

        if vrsta_meritve == "Z auto":
            with open(
                CSVFILE_INSTALACIJE_OSNOVNE,
                "a",
                encoding="utf-8",
                newline="",
            ) as csvfile:
                writer = csv.writer(
                    csvfile, delimiter=";", quotechar='"', quoting=csv.QUOTE_MINIMAL
                )

                tip_varovalke = meritev.najdi_tip_varovalke()
                I_varovalke = meritev.najdi_I_varovalke()
                t_varovalke = meritev.najdi_t_varovalke()
                isc_faktor = meritev.najdi_Isc_faktor()
                zln = meritev.najdi_Z_LN()
                ipsc_ln = meritev.najdi_Ipsc_LN()
                ipsc_lpe = meritev.najdi_Ipsc_LPE()
                dU = meritev.najdi_dU()
                zlpe = meritev.najdi_Z_LPE()
                komentar = meritev.najdi_komentar()

                st_vnesenih_meritev += 1
                array_ki_ga_zapisemo_v_csv = [
                    st_vnesenih_meritev,
                    ime,
                    PRAZNO,
                    PRAZNO,
                    PRAZNO,
                    glavna_izenac_povezava,
                    PRAZNO,
                    rlpe,
                    tip_varovalke,
                    I_varovalke,
                    t_varovalke,
                    f"{zlpe}/{ipsc_ln}",
                    f"{zln}/{ipsc_lpe}/{dU}",
                    PRAZNO,
                    I_dN,
                    PRAZNO,
                    t1x,
                    t5x,
                    Uc,
                    PRAZNO,
                    komentar,
                    vrsta_meritve,
                    uln,
                    maxRplusRminus,
                    isc_faktor,
                    ia_psc_navidezni_stolpec,
                    pot,
                ]
                writer.writerow(array_ki_ga_zapisemo_v_csv)
                csvfile.close()

    if (
        "Zloop" in vrste_meritev_v_kocki
        or "Z LINE" in vrste_meritev_v_kocki
        or "Z loop 4W" in vrste_meritev_v_kocki
        or "Z line 4W" in vrste_meritev_v_kocki
    ):
        dodaten4W = (
            "Z loop 4W" in vrste_meritev_v_kocki or "Z line 4W" in vrste_meritev_v_kocki
        )
        st_zline = 0
        st_zloop = 0
        ustrezni_zloop = []
        ustrezni_zline_230 = []
        ustrezni_zline_400 = []
        for meritev in kocka:
            vrsta_meritve = meritev.doloci_vrsto_meritve()
            if vrsta_meritve in ["Zloop", "Z loop 4W"]:
                st_zloop += 1
                ustrezni_zloop.append(meritev)

            if vrsta_meritve in ["Z LINE", "Z line 4W"]:
                st_zline += 1
                if dodaten4W:
                    if (
                        meritev.najdi_Uln() != "X"
                        and float(meritev.najdi_Uln()) > 200
                        and float(meritev.najdi_Uln()) < 270
                    ):
                        ustrezni_zline_230.append(meritev)
                    else:
                        ustrezni_zline_400.append(meritev)
                else:
                    if "230" == meritev.najdi_Un():
                        ustrezni_zline_230.append(meritev)
                    else:
                        ustrezni_zline_400.append(meritev)

                # if "400" == meritev.najdi_Un():
                #     ustrezni_zline_3.append(meritev)

        # Safety check v primeru, da ni treh zloop/zline elementov
        if (
            # len(ustrezni_zline_400) != len(ustrezni_zloop)
            # or len(ustrezni_zline_400) != 3
            # or len(ustrezni_zloop) != 3
            # Tole morda spremeniti
        ):
            print("\nNapaka: Nekaj ni v redu s številom zloop/zlinov")
            print("Pot problematične meritve:", pot.strip())
            print("Dolžina zloop", len(ustrezni_zloop))
            print("Dolžina zline", len(ustrezni_zline_400))

        else:
            ustrezni_zline_pravi = []
            if ustrezni_zline_230:
                ustrezni_zline_pravi = ustrezni_zline_230
                while len(ustrezni_zline_230) > len(ustrezni_zloop):
                    ustrezni_zloop.append(Meritev(""))
                while len(ustrezni_zline_230) < len(ustrezni_zloop):
                    ustrezni_zline_230.append(Meritev(""))

            elif ustrezni_zline_400:
                ustrezni_zline_pravi = ustrezni_zline_400
                while len(ustrezni_zline_400) > len(ustrezni_zloop):
                    ustrezni_zloop.append(Meritev(""))
                while len(ustrezni_zline_400) < len(ustrezni_zloop):
                    ustrezni_zline_400.append(Meritev(""))

            for i in range(len(ustrezni_zloop)):
                with open(
                    CSVFILE_INSTALACIJE_OSNOVNE,
                    "a",
                    encoding="utf-8",
                    newline="",
                ) as csvfile:
                    writer = csv.writer(
                        csvfile,
                        delimiter=";",
                        quotechar='"',
                        quoting=csv.QUOTE_MINIMAL,
                    )

                    if ustrezni_zline_pravi:
                        if ustrezni_zline_pravi[i].besedilo.count("p//") > 0:
                            ipsc_zline, z_zline = "X", "X"
                        else:
                            ipsc_zline = ustrezni_zline_pravi[i].najdi_Ipsc()
                            z_zline = ustrezni_zline_pravi[i].najdi_Z()

                        if ustrezni_zloop[i].besedilo.count("p//") > 0:
                            ipsc_zloop, z_zloop = "X", "X"
                        else:
                            ipsc_zloop = ustrezni_zloop[i].najdi_Ipsc()
                            z_zloop = ustrezni_zloop[i].najdi_Z()

                        if ustrezni_zloop[i].besedilo.count("p//") > 0:
                            (
                                uln,
                                ipsc_lpe,
                                zlpe,
                                ia_psc_navidezni_stolpec,
                                tip_varovalke,
                                I_varovalke,
                                t_varovalke,
                                isc_faktor,
                                komentar,
                            ) = ("X" for _ in range(9))
                        else:
                            uln = (
                                ustrezni_zline_pravi[i].najdi_Uln()
                                if dodaten4W
                                else ustrezni_zline_pravi[i].najdi_Un()
                            )
                            ipsc_lpe = ustrezni_zloop[i].najdi_Ipsc_LPE()
                            zlpe = ustrezni_zloop[i].najdi_Z_LPE()
                            ia_psc_navidezni_stolpec = ustrezni_zloop[i].najdi_Ia_Ipsc()
                            tip_varovalke = ustrezni_zloop[i].najdi_tip_varovalke()
                            I_varovalke = ustrezni_zloop[i].najdi_I_varovalke()
                            t_varovalke = ustrezni_zloop[i].najdi_t_varovalke()
                            isc_faktor = ustrezni_zloop[i].najdi_Isc_faktor()
                            vrsta_meritve = (
                                "ZLOOP / ZLINE"
                                if not dodaten4W
                                else "ZLOOP 4W / ZLINE 4W"
                            )
                            komentar = ustrezni_zloop[i].najdi_komentar()

                            if not dU:
                                dU = "X"

                            st_vnesenih_meritev += 1
                            array_ki_ga_zapisemo_v_csv = [
                                st_vnesenih_meritev,
                                ime,
                                PRAZNO,
                                PRAZNO,
                                PRAZNO,
                                glavna_izenac_povezava,
                                PRAZNO,
                                rlpe,
                                tip_varovalke,
                                I_varovalke,
                                t_varovalke,
                                f"{z_zloop}/{ipsc_zloop}",
                                f"{z_zline}/{ipsc_zline}/{dU}",
                                PRAZNO,
                                I_dN,
                                PRAZNO,
                                t1x,
                                t5x,
                                Uc,
                                PRAZNO,
                                komentar,
                                vrsta_meritve,
                                uln,
                                maxRplusRminus,
                                isc_faktor,
                                ia_psc_navidezni_stolpec,
                                pot,
                            ]
                            writer.writerow(array_ki_ga_zapisemo_v_csv)
                            csvfile.close()

                    else:
                        (
                            uln,
                            ipsc_lpe,
                            zlpe,
                            ia_psc_navidezni_stolpec,
                            tip_varovalke,
                            I_varovalke,
                            t_varovalke,
                            isc_faktor,
                            komentar,
                        ) = ("X" for _ in range(9))

    for meritev in kocka:
        vrsta_meritve = meritev.doloci_vrsto_meritve()
        if vrsta_meritve == "RCD Auto":
            with open(
                CSVFILE_INSTALACIJE_RCD,
                "a",
                encoding="utf-8",
                newline="",
            ) as csvfile_RCD:
                writer = csv.writer(
                    csvfile_RCD, delimiter=";", quotechar='"', quoting=csv.QUOTE_MINIMAL
                )
                I_dN = meritev.najdi_I_dN()
                tip_rcd = meritev.najdi_Tip()
                Uc = meritev.najdi_Uc()
                ozemljitveni_sistem = meritev.najdi_Ozemljitveni_sistem()
                komentar = meritev.najdi_komentar()

                t1x_plus = pretvori_string_milisekund_v_ustrezen_format(
                    meritev.najdi_t_IΔN_x1_plus()
                )
                t1x_neg = pretvori_string_milisekund_v_ustrezen_format(
                    meritev.najdi_t_IΔN_x1_minus()
                )
                t5x_plus = pretvori_string_milisekund_v_ustrezen_format(
                    meritev.najdi_t_IΔN_x5_plus()
                )
                t5x_neg = pretvori_string_milisekund_v_ustrezen_format(
                    meritev.najdi_t_IΔN_x5_minus()
                )

                # V primeru, da vse količine obstajajo
                if "X" not in [t1x_plus, t1x_neg, t5x_plus, t5x_neg]:
                    t1x_plus = float(t1x_plus)
                    t1x_neg = float(t1x_neg)
                    t5x_plus = float(t5x_plus)
                    t5x_neg = float(t5x_neg)

                    # Pri kateri velikosti spremenimo v int?
                    if max(t1x_plus, t1x_neg) >= 100:
                        t1x = f"{int(max(t1x_plus, t1x_neg))}".replace(".", ",")
                    else:
                        t1x = f"{max(t1x_plus, t1x_neg)}".replace(".", ",")

                    if max(t5x_plus, t5x_neg) >= 100:
                        t5x = f"{int(max(t5x_plus, t5x_neg))}".replace(".", ",")
                    else:
                        t5x = f"{max(t5x_plus, t5x_neg)}".replace(".", ",")
                else:
                    print("\nNapaka: manjkajoči podatki t1x ali t5x pri", pot.strip())
                    t1x = "X"
                    t5x = "X"

                st_vnesenih_meritev_RCD += 1
                array_ki_ga_zapisemo_v_csv = [
                    st_vnesenih_meritev_RCD,
                    ime,
                    PRAZNO,
                    PRAZNO,
                    I_dN,
                    tip_rcd,
                    PRAZNO,
                    PRAZNO,
                    PRAZNO,
                    PRAZNO,
                    Uc,
                    PRAZNO,
                    t1x,
                    t5x,
                    ozemljitveni_sistem,
                    komentar,
                    vrsta_meritve,
                    pot,
                ]
                writer.writerow(array_ki_ga_zapisemo_v_csv)
                csvfile_RCD.close()


def zapisi_kocko_meritev_v_excel_stroji(
    kocka,
    loceno_besedilo,
    slovar_kock_in_ustreznih_poti,
    t_varovalke_neprekinjenost,
    I_varovalke_neprekinjenost,
    tip_varovalke_neprekinjenost,
    meja_izolacijske_upornosti_stroji_riso_rdeca,
    prevedi_v_anglescino,
    napetost_dotika,
):
    (
        komentar,
        pot,
        pot_locena_na_elemente,
        ime,
        vrste_meritev_v_kocki,
        slovar_vrst_meritev,
    ) = pridobi_potne_elemente(kocka, loceno_besedilo, slovar_kock_in_ustreznih_poti)

    if slovar_vrst_meritev["Zloop"] > 0:
        with open(CSVFILE_STROJI_ZLOOP, "a", encoding="utf-8", newline="") as csvfile:
            writer = csv.writer(
                csvfile, delimiter=";", quotechar='"', quoting=csv.QUOTE_MINIMAL
            )

            for meritev in kocka:
                if meritev.doloci_vrsto_meritve() == "Zloop":
                    komentar = prevedi_s_slovarjem(
                        meritev.najdi_komentar(), prevedi_v_anglescino
                    )

                    tip_varovalke = meritev.najdi_tip_varovalke()
                    t_varovalke = (
                        "X"
                        if meritev.najdi_t_varovalke() == "X"
                        else float(
                            meritev.najdi_t_varovalke()
                            .replace(" s", "")
                            .replace(",", ".")
                        )
                    )
                    i_varovalke = (
                        "X"
                        if meritev.najdi_I_varovalke() == "X"
                        else float(
                            meritev.najdi_I_varovalke()
                            .replace(" A", "")
                            .replace(",", ".")
                        )
                    )

                    if i_varovalke == "X":
                        array_ki_ga_zapisemo_v_csv = ["X" for _ in range(11)] + [pot]
                        writer.writerow(array_ki_ga_zapisemo_v_csv)
                    else:
                        Un = int(meritev.najdi_Un().replace(" V", ""))
                        Ipsc = float(
                            meritev.najdi_Ia_Ipsc().replace(",", ".").replace(">", "")
                        )
                        Z = (
                            "X"
                            if meritev.najdi_Z() == "X"
                            else float(
                                meritev.najdi_Z().replace(",", ".").replace(">", "")
                            )
                        )

                        excel_delovna_datoteka = load_workbook(
                            MEJE_ZA_MERITVE, data_only=True
                        )

                        excel_delovni_list = excel_delovna_datoteka["gG"]
                        prva_vrstica = 6
                        zadnja_vrstica = 34

                        t_varovalke_je_ustrezen = False
                        for vrednost in [0.1, 0.2, 0.4, 5.0]:
                            if t_varovalke == vrednost:
                                t_varovalke_je_ustrezen = True
                                break

                        if not t_varovalke_je_ustrezen:
                            tok_zascite = "X"
                            izracun = "X"

                        slovar_t_varovalk_in_stolpcev = {
                            0.1: 1,
                            0.2: 4,
                            0.4: 7,
                            5.0: 10,
                        }
                        if t_varovalke_neprekinjenost in slovar_t_varovalk_in_stolpcev:
                            stolpec = slovar_t_varovalk_in_stolpcev[
                                t_varovalke_neprekinjenost
                            ]
                        else:
                            print("Nemogoče trajanje")

                        if t_varovalke == 0.1:
                            zadnja_vrstica = 30

                        stolpec_0 = [
                            excel_delovni_list.cell(row=i, column=stolpec).value
                            for i in range(prva_vrstica, zadnja_vrstica + 1)
                        ]

                        tok_zascite = excel_delovni_list.cell(
                            row=stolpec_0.index(i_varovalke) + 6, column=stolpec + 1
                        ).value
                        izracun = 2 / 3 * (Un / tok_zascite)

                        excel_delovni_list = excel_delovna_datoteka["ZLOOP"]
                        prva_vrstica = 2
                        zadnja_vrstica = 13

                        stolpec_1 = [
                            excel_delovni_list.cell(row=i, column=1).value
                            for i in range(prva_vrstica, zadnja_vrstica + 1)
                        ]

                        stolpec_2 = [
                            excel_delovni_list.cell(row=i, column=2).value
                            for i in range(prva_vrstica, zadnja_vrstica + 1)
                        ]

                        stolpec_3 = [
                            excel_delovni_list.cell(row=i, column=3).value
                            for i in range(prva_vrstica, zadnja_vrstica + 1)
                        ]

                        # to se načeloma ne bi smelo zgoditi
                        if i_varovalke not in stolpec_3:
                            tok_zascite = "X"
                            izracun = "X"
                        else:
                            idx = stolpec_3.index(i_varovalke)
                            Z_meja = stolpec_1[idx]
                            min_vodnik = stolpec_2[idx]

                            slovar_tipov_varovalk_in_stolpcev = {
                                "NV": 4,
                                "gG": 4,
                                "gL": 4,
                                "B": 6,
                                "C": 7,
                                "D": 8,
                            }

                            if tip_varovalke not in slovar_tipov_varovalk_in_stolpcev:
                                max_meter = "X"

                            if t_varovalke not in [0.4, 5.0]:
                                max_meter = "X"
                            dodaten_zamik_stolpca = 1 if t_varovalke == 0.4 else 0

                            stolpec_metra = [
                                excel_delovni_list.cell(
                                    row=i,
                                    column=slovar_tipov_varovalk_in_stolpcev[
                                        tip_varovalke
                                    ]
                                    + dodaten_zamik_stolpca,
                                ).value
                                for i in range(prva_vrstica, zadnja_vrstica + 1)
                            ]

                            max_meter = stolpec_metra[idx]
                            oknok = (
                                "✓"
                                if (
                                    (int(Z_meja) + 25) / 1000 > Z and tok_zascite < Ipsc
                                )
                                else "✗"
                            )

                            array_ki_ga_zapisemo_v_csv = [
                                PRAZNO,
                                komentar,
                                t_varovalke,
                                Un,
                                tok_zascite,
                                str(izracun).replace(".", ","),
                                f"{Ipsc}/{str(Z)}".replace(".", ","),
                                oknok,
                                min_vodnik,
                                max_meter,
                                Z_meja,
                                pot,
                            ]
                            writer.writerow(array_ki_ga_zapisemo_v_csv)
            csvfile.close()

    if slovar_vrst_meritev["R iso"] + slovar_vrst_meritev["R IZO"] > 0:
        with open(CSVFILE_STROJI_RISO, "a", encoding="utf-8", newline="") as csvfile:
            writer = csv.writer(
                csvfile, delimiter=";", quotechar='"', quoting=csv.QUOTE_MINIMAL
            )

            for meritev in kocka:
                if meritev.doloci_vrsto_meritve() in ["R iso", "R IZO"]:
                    komentar = meritev.najdi_komentar()
                    riso = meritev.najdi_Riso()

                    krizec_kljukica = (
                        "X"
                        if riso == "X"
                        else (
                            "✗"
                            if float(
                                riso.replace(">", "")
                                .replace(",", ".")
                                .replace(" MΩ", "")
                            )
                            < meja_izolacijske_upornosti_stroji_riso_rdeca
                            else "✓"
                        )
                    )

                    array_ki_ga_zapisemo_v_csv = [
                        PRAZNO,
                        PRAZNO,
                        PRAZNO,
                        riso,
                        krizec_kljukica,
                        komentar,
                    ]
                    writer.writerow(array_ki_ga_zapisemo_v_csv)
            csvfile.close()

    if (
        slovar_vrst_meritev["Discharge time"] + slovar_vrst_meritev["Čas praznjenja"]
        > 0
    ):
        with open(
            CSVFILE_STROJI_DISCHARGE_TIME, "a", encoding="utf-8", newline=""
        ) as csvfile:
            writer = csv.writer(
                csvfile, delimiter=";", quotechar='"', quoting=csv.QUOTE_MINIMAL
            )

            for meritev in kocka:
                if meritev.doloci_vrsto_meritve() in [
                    "Discharge time",
                    "Čas praznjenja",
                ]:
                    komentar = meritev.najdi_komentar()

                    t = float(meritev.najdi_t().replace(",", "."))
                    meja_t = meritev.najdi_meja_t()

                    if meja_t == "5":
                        drugi_tretji = [t, PRAZNO]
                    elif meja_t == "1":
                        drugi_tretji = [PRAZNO, t]
                    else:
                        drugi_tretji = ["X", "X"]
                        print("meja_t ni niti 1s ali 5s")

                    array_ki_ga_zapisemo_v_csv = [PRAZNO, *drugi_tretji, komentar]
                    writer.writerow(array_ki_ga_zapisemo_v_csv)
            csvfile.close()

    if slovar_vrst_meritev["Neprekinjenost"] > 0:
        with open(
            CSVFILE_STROJI_NEPREKINJENOST, "a", encoding="utf-8", newline=""
        ) as csvfile:
            writer = csv.writer(
                csvfile, delimiter=";", quotechar='"', quoting=csv.QUOTE_MINIMAL
            )

            # To se da verjetno nekoliko olepšati, ker se od 14. 4. 2024 gleda Neprekinjenost in R low 4 ločeno.

            for meritev in kocka:
                if meritev.doloci_vrsto_meritve() in ["Neprekinjenost"]:
                    # if meritev.doloci_vrsto_meritve() == "R low 4":
                    #     # R low v resnici vsebuje vse te atribute
                    #     t_varovalke_neprekinjenost = meritev.najdi_t_varovalke()
                    #     I_varovalke_neprekinjenost = meritev.najdi_I_varovalke()
                    #     tip_varovalke_neprekinjenost = meritev.najdi_tip_varovalke()

                    komentar = meritev.najdi_komentar()
                    R = (
                        "X"
                        if meritev.najdi_R() == "X"
                        else float(meritev.najdi_R().replace(",", ".").replace(">", ""))
                    )

                    trajanje = (
                        "X"
                        if meritev.najdi_trajanje() == "X"
                        else float(
                            meritev.najdi_trajanje().replace(" s", "").replace(",", ".")
                        )
                    )
                    i_out = (
                        "X"
                        if meritev.najdi_I_out() == "X"
                        else float(
                            meritev.najdi_I_out().replace(" A", "").replace(",", ".")
                        )
                    )

                    t_varovalke_je_ustrezen = False
                    for vrednost in [0.1, 0.2, 0.4, 5.0]:
                        if t_varovalke_neprekinjenost == vrednost:
                            t_varovalke_je_ustrezen = True
                            break

                    if not t_varovalke_je_ustrezen:
                        tok_zascite = "X"
                        izracun = "X"

                    excel_delovna_datoteka = load_workbook(
                        MEJE_ZA_MERITVE, data_only=True
                    )

                    if tip_varovalke_neprekinjenost in ["gG", "gL", "NV"]:
                        excel_delovni_list = excel_delovna_datoteka["gG"]
                        prva_vrstica = 6
                        zadnja_vrstica = 34

                        slovar_t_varovalk_in_stolpcev = {
                            0.1: 1,
                            0.2: 4,
                            0.4: 7,
                            5.0: 10,
                        }
                        if t_varovalke_neprekinjenost in slovar_t_varovalk_in_stolpcev:
                            stolpec = slovar_t_varovalk_in_stolpcev[
                                t_varovalke_neprekinjenost
                            ]
                        else:
                            print("Nemogoče trajanje")

                        if t_varovalke_neprekinjenost == 0.1:
                            zadnja_vrstica = 30

                        stolpec_1 = [
                            excel_delovni_list.cell(row=i, column=stolpec).value
                            for i in range(prva_vrstica, zadnja_vrstica + 1)
                        ]

                        if i_out not in stolpec_1:
                            tok_zascite = "X"
                            izracun = "X"
                        else:
                            tok_zascite = excel_delovni_list.cell(
                                row=stolpec_1.index(i_out) + 6, column=stolpec + 1
                            ).value
                            izracun = min((tok_zascite / napetost_dotika), 0.3)
                            krizec_kljukica = (
                                "X" if R == "X" else ("✓" if izracun > R else "✗")
                            )

                        array_ki_ga_zapisemo_v_csv = [
                            PRAZNO,
                            komentar,
                            i_out,
                            R,
                            PRAZNO,
                            krizec_kljukica,
                            trajanje,
                            pot,
                        ]
                        writer.writerow(array_ki_ga_zapisemo_v_csv)

                    if tip_varovalke_neprekinjenost in ["B", "C", "D", "K"]:
                        excel_delovni_list = excel_delovna_datoteka["BCD"]
                        prva_vrstica = 6
                        zadnja_vrstica = 18

                        if tip_varovalke_neprekinjenost == "K":
                            stolpec_K = [
                                excel_delovni_list.cell(row=i, column=8).value
                                for i in range(prva_vrstica, zadnja_vrstica + 1)
                            ]

                            stolpec_K_drugi = [
                                excel_delovni_list.cell(row=i, column=9).value
                                for i in range(prva_vrstica, zadnja_vrstica + 1)
                            ]

                            tok_zascite = stolpec_K_drugi[
                                stolpec_K.index(I_varovalke_neprekinjenost)
                            ]

                        else:
                            slovar_tipa_varovalk_in_stolpcev = {
                                "B": 2,
                                "C": 4,
                                "D": 6,
                            }

                            stolpec = slovar_tipa_varovalk_in_stolpcev[
                                tip_varovalke_neprekinjenost
                            ]

                            stolpec_1 = [
                                excel_delovni_list.cell(row=i, column=1).value
                                for i in range(prva_vrstica, zadnja_vrstica + 1)
                            ]

                            stolpec_2 = [
                                excel_delovni_list.cell(row=i, column=stolpec).value
                                for i in range(prva_vrstica, zadnja_vrstica + 1)
                            ]

                            tok_zascite = stolpec_2[
                                stolpec_1.index(int(I_varovalke_neprekinjenost))
                            ]

                        izracun = min((tok_zascite / napetost_dotika), 0.3)
                        print(type(R), R)
                        krizec_kljukica = "✓" if R != "X" and izracun > R else "✗"
                        array_ki_ga_zapisemo_v_csv = [
                            PRAZNO,
                            komentar,
                            i_out,
                            R,
                            PRAZNO,
                            krizec_kljukica,
                            trajanje,
                            pot,
                        ]
                        writer.writerow(array_ki_ga_zapisemo_v_csv)
            csvfile.close()

    if slovar_vrst_meritev["R low 4"] > 0:
        with open(CSVFILE_STROJI_RLOW4, "a", encoding="utf-8", newline="") as csvfile:
            writer = csv.writer(
                csvfile, delimiter=";", quotechar='"', quoting=csv.QUOTE_MINIMAL
            )
            # To sem pustil, ker bo morda kdaj uporabno?

            # seznam_Rjev_za_rlow4_meritve = []

            # Spodnji loop se da močno izboljšati, veliko stvari je nekoliko nenavadnih

            for meritev in kocka:
                if meritev.doloci_vrsto_meritve() == "R low 4":
                    komentar = meritev.najdi_komentar()
                    # seznam_Rjev_za_rlow4_meritve.append(meritev.najdi_R())
                    if (
                        # >1999 je zgornja meja, nastavljena na napravi
                        ">1999" in meritev.najdi_R_pozitivno()
                        or ">1999" in meritev.najdi_R_negativno()
                    ):
                        maxRplusRminus = ">1999"
                    else:
                        R_pozitivno_int = (
                            meritev.najdi_R_pozitivno()
                            .replace(",", ".")
                            .replace(" Ω", "")
                            .replace(">", "")
                        )
                        if R_pozitivno_int != "X":
                            R_pozitivno_int = float(R_pozitivno_int)

                        R_negativno_int = (
                            meritev.najdi_R_negativno()
                            .replace(",", ".")
                            .replace(" Ω", "")
                            .replace(">", "")
                        )
                        if R_negativno_int != "X":
                            R_negativno_int = float(R_negativno_int)

                        # Pri kateri velikosti spremenimo v int?
                        if R_negativno_int == "X" or R_pozitivno_int == "X":
                            maxRplusRminus = "X"
                        else:
                            if max(R_negativno_int, R_pozitivno_int) >= 100:
                                maxRplusRminus = (
                                    f"{int(max(R_negativno_int, R_pozitivno_int))}"
                                )
                            else:
                                maxRplusRminus = (
                                    f"{max(R_negativno_int, R_pozitivno_int)}"
                                )
                    R = meritev.najdi_R()
                    array_ki_ga_zapisemo_v_csv = [
                        ime,
                        R,
                        maxRplusRminus,
                        komentar,
                        pot,
                    ]
                    writer.writerow(array_ki_ga_zapisemo_v_csv)
            csvfile.close()
            # Tole dvoje izgleda precej neuporabno
            # seznam_Rjev_za_rlow4_meritve.sort(key=lambda x: vrednoti_string(x))
            # rlow4_meritev_z_minimalno = seznam_Rjev_za_rlow4_meritve[0]


def zapisi_kocko_meritev_v_excel_elektricne_omare(
    kocka, loceno_besedilo, slovar_kock_in_ustreznih_poti, prevedi_v_anglescino
):
    (
        komentar,
        pot,
        pot_locena_na_elemente,
        ime,
        vrste_meritev_v_kocki,
        slovar_vrst_meritev,
    ) = pridobi_potne_elemente(kocka, loceno_besedilo, slovar_kock_in_ustreznih_poti)

    if "Page" in komentar:
        komentar = komentar[0 : komentar.index("Page")]

    komentar = oklesti_komentar(komentar, "Page")

    if slovar_vrst_meritev["R low 4"] + slovar_vrst_meritev["Neprekinjenost"] > 0:
        with open(
            CSVFILE_ELEKTRICNE_OMARE, "a", encoding="utf-8", newline=""
        ) as csvfile:
            writer = csv.writer(
                csvfile, delimiter=";", quotechar='"', quoting=csv.QUOTE_MINIMAL
            )
            for meritev in kocka:
                komentar = prevedi_s_slovarjem(
                    meritev.najdi_komentar(), prevedi_v_anglescino=prevedi_v_anglescino
                )
                R = meritev.najdi_R()
                if R == "X":
                    oknok = "X"
                else:
                    oknok = "OK" if vrednoti_string(R) < 0.1 else "NOK"

                array_ki_ga_zapisemo_v_csv = [komentar, R, oknok, pot]
                writer.writerow(array_ki_ga_zapisemo_v_csv)
        csvfile.close()


def zapisi_kocko_meritev_v_excel_strelovodi(
    kocka, loceno_besedilo, slovar_kock_in_ustreznih_poti
):
    # Tukaj je že samo po sebi veliko neporabljenih elementov
    (
        komentar,
        pot,
        pot_locena_na_elemente,
        ime,
        vrste_meritev_v_kocki,
        slovar_vrst_meritev,
    ) = pridobi_potne_elemente(kocka, loceno_besedilo, slovar_kock_in_ustreznih_poti)

    if slovar_vrst_meritev["R low 4"] > 0:
        with open(CSVFILE_STRELOVODI, "a", encoding="utf-8", newline="") as csvfile:
            writer = csv.writer(
                csvfile, delimiter=";", quotechar='"', quoting=csv.QUOTE_MINIMAL
            )
            for meritev in kocka:
                komentar = meritev.najdi_komentar()
                R = meritev.najdi_R()
                dane = "DA" if vrednoti_string(R) < 10 else "NE"

                vzorec = r"MS\d+"
                prva_vrstica = (
                    re.search(vzorec, komentar) if "MS" in komentar else PRAZNO
                )

                array_ki_ga_zapisemo_v_csv = [
                    prva_vrstica,
                    komentar,
                    R,
                    PRAZNO,
                    PRAZNO,
                    PRAZNO,
                    PRAZNO,
                    dane,
                    pot,
                ]
                writer.writerow(array_ki_ga_zapisemo_v_csv)
        csvfile.close()

    if slovar_vrst_meritev["Ozemljitvena upornost"] > 0:
        with open(CSVFILE_STRELOVODI, "a", encoding="utf-8", newline="") as csvfile:
            writer = csv.writer(
                csvfile, delimiter=";", quotechar='"', quoting=csv.QUOTE_MINIMAL
            )
            for meritev in kocka:
                komentar = meritev.najdi_komentar()
                Re = meritev.najdi_Re()
                if Re == "X":
                    dane = "X"
                else:
                    dane = "DA" if float(Re.replace(",", ".")) < 10 else "NE"

                vzorec = r"MS\d+"
                prva_vrstica = (
                    re.search(vzorec, komentar).group() if "MS" in komentar else PRAZNO
                )

                array_ki_ga_zapisemo_v_csv = [
                    prva_vrstica,
                    komentar,
                    Re,
                    PRAZNO,
                    PRAZNO,
                    PRAZNO,
                    PRAZNO,
                    dane,
                    pot,
                ]
                writer.writerow(array_ki_ga_zapisemo_v_csv)
        csvfile.close()


def najdi_po_vrsti_urejen_seznam_datumov(vse_besedilo):
    loceno_besedilo_po_presledkih = vse_besedilo.split()
    seznam_stringov_z_datumi = []
    for i in loceno_besedilo_po_presledkih:
        if re.search(r"\d{2}\.\d{2}\.\d{4}", i):
            # vsi datumi so pravilne oblike (zraven so zapisane odvečne ničle),
            # zato je upravičeno tole
            string_datuma = re.search(r"\d{2}\.\d{2}\.\d{4}", i).group()
            datum = datetime.strptime(string_datuma, "%d.%m.%Y")
            seznam_stringov_z_datumi.append(datum)
    return [
        datetime.strftime(datum, "%d.%m.%Y")
        for datum in sorted(seznam_stringov_z_datumi)
    ]


# Pot se obnaša nekoliko drugače kot ostale komponente, zato je izven
# Tole se morda da izboljšati z regexom
def najdi_pot_izven_razreda_Meritev(besedilo):
    idx = besedilo.find("Pot:")
    string_ki_ga_obdelujemo = besedilo[idx:]
    if "Page" in string_ki_ga_obdelujemo:
        idx = string_ki_ga_obdelujemo.find("Page")
        # poanta je v temu, da so spredaj odvečne črke
        string_ki_ga_obdelujemo = string_ki_ga_obdelujemo[:idx]
    if "Serijsko" in string_ki_ga_obdelujemo:
        idx = string_ki_ga_obdelujemo.find("Serijsko")
        string_ki_ga_obdelujemo = string_ki_ga_obdelujemo[:idx]
    return string_ki_ga_obdelujemo
